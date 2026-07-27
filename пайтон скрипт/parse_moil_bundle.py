#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pdfplumber
from openpyxl import load_workbook


INVOICE_NO_RE = re.compile(r'Invoice No\.?\s*[:#]?\s*([0-9A-Za-z\-/]+)', re.I)
PACKAGE_HEADER_RE_OLD = re.compile(r'^\s*(\d+)\s+\d+\s+PL\b')
PACKAGE_HEADER_RE_NEW = re.compile(r'^\s*(\d+)\s+(?:PL|CB)\d+\s+(?:PL|CB)\b', re.I)
FNO_PACKAGE_HEADER_RE = re.compile(
    r'^\s*(\d+)\s+(\d+)\s+(PL|CB)\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s*$',
    re.I,
)
IN_CB_RE = re.compile(r'IN\s+CB\s+(\d+)', re.I)
PACKING_ROW_RE_OLD = re.compile(r'^([A-Z][A-Z0-9-]*)\s+(.+?)\s+(\d[\d,]*\.\d+)\s+(\d[\d,]*\.\d+)\s+(\d[\d,]*\.?\d*)\s*$')
PACKING_ROW_RE_NEW = re.compile(r'^([A-Z][A-Z0-9-]*)\s+(.+?)\s+(\d[\d,]*\.\d+)\s+(\d[\d,]*\.\d+)\s+(\d[\d,]*\.\d+)\s*$')
FNO_PACKING_ROW_RE = re.compile(
    r'^([A-Z][A-Z0-9-]*)\s+(.+?)\s+'
    r'(?:(\d{8,14})\s+)?'
    r'(\d[\d,]*(?:\.\d+)?)\s+'
    r'(?:(\d[\d,]*(?:\.\d+)?)\s+)?'
    r'(\d[\d,]*(?:\.\d+)?)\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s*$'
)
FNO_INVOICE_ROW_RE = re.compile(
    r'^(\d{1,3})\s+'
    r'(?:(FOC)\s+)?'
    r'([A-Z0-9-]+)\s+'
    r'(.+?)\s+'
    r'(\d[\d,]*)\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s+\$\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s+\$\s+'
    r'(\d+(?:\.\d+)?)%\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s+\$\s+'
    r'(\d[\d,]*(?:\.\d+)?)\s+\$\s+'
    r'([A-Za-z]+)\s*$'
)
INVOICE_DATA_RE = re.compile(
    r'^([A-Z0-9-]+)'
    r'(?:\s+(.+?))?'
    r'\s+(\d[\d,]*)'
    r'\s+(\d[\d,]*\.\d+)'
    r'\s+\$\s+(\d[\d,]*\.\d+)'
    r'\s+\$\s+(\d+(?:\.\d+)?)%'
    r'\s+(\d[\d,]*\.\d+)'
    r'\s+\$\s+(\d[\d,]*\.\d+)'
    r'\s+\$\s+(\d[\d,]*\.\d+)'
    r'\s+\$\s+([A-Za-z]+)$'
)
INVOICE_ROW_SINGLE_RE = re.compile(
    r'^([A-Z0-9-]+)\s+(\d+)\s+(\d[\d,]*)\s+(\d+(?:\.\d+)?)%'
    r'\s+(.+?)\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+'
    r'(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+([A-Za-z]+)$'
)
INVOICE_ROW_SINGLE_RE_NOT_APPLICABLE = re.compile(
    r'^([A-Z0-9-]+)\s+(\d+)\s+(\d[\d,]*)\s+(\d+(?:\.\d+)?)%'
    r'\s+(.+?)\s+(Not Applicable)\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+'
    r'(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$$'
)
MANIFEST_FILE_NAMES = ('manifest.json', 'shipment_manifest.json')


def normalize_space(value: Any) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_sku(value: Any) -> str:
    return normalize_space(value).upper()


def normalize_header(value: Any) -> str:
    return re.sub(r'[^a-z0-9а-яё]+', '', str(value or '').lower())


def parse_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    text = normalize_space(value).replace('$', '').replace('%', '').replace(' ', '').replace(',', '')
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def dec_to_num(value: Optional[Decimal]) -> Optional[float]:
    if value is None:
        return None
    return float(value)


def load_manifest(input_dir: str) -> Tuple[Optional[str], Optional[Dict[str, Any]]]:
    for file_name in MANIFEST_FILE_NAMES:
        path = os.path.join(input_dir, file_name)
        if not os.path.exists(path):
            continue
        try:
            with open(path, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
            if isinstance(data, dict):
                return path, data
        except Exception:
            pass
    return None, None


def infer_file_role(file_name: str, meta: Optional[Dict[str, Any]] = None) -> str:
    meta = meta or {}
    doc_type = normalize_space(meta.get('doc_type') or meta.get('file_type')).lower()
    if doc_type in {'inv', 'invoice'}:
        return 'inv'
    if doc_type in {'pac', 'packing'}:
        return 'pac'
    if doc_type == 'batch':
        return 'batch'

    low = file_name.lower()
    if re.search(r'(^|[-_])inv([-. _]|$)', low):
        return 'inv'
    if re.search(r'(^|[-_])pac([-. _]|$)', low):
        return 'pac'
    if 'batch' in low:
        return 'batch'
    return 'unknown'


def build_manifest_entries(input_dir: str, manifest: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not isinstance(manifest, dict):
        return []
    result: List[Dict[str, Any]] = []
    for raw in manifest.get('files') or []:
        if not isinstance(raw, dict):
            continue
        saved_path = normalize_space(raw.get('saved_path'))
        saved_name = normalize_space(raw.get('saved_name'))
        path = saved_path or (os.path.join(input_dir, saved_name) if saved_name else '')
        if not path or not os.path.exists(path):
            continue
        result.append({
            'path': path,
            'saved_name': saved_name or os.path.basename(path),
            'original_name': normalize_space(raw.get('original_name')) or None,
            'doc_type': normalize_space(raw.get('doc_type')) or None,
            'meta': raw,
        })
    return result


def extract_pdf_lines(pdf_path: str) -> List[str]:
    result: List[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=2) or ''
            for raw in text.splitlines():
                line = normalize_space(raw)
                if line:
                    result.append(line)
    return result


def extract_invoice_no(lines: Sequence[str], fallback_file_name: str) -> str:
    full_text = '\n'.join(lines)
    match = INVOICE_NO_RE.search(full_text)
    if match:
        return normalize_space(match.group(1))
    m = re.search(r'(\d{6,})', fallback_file_name)
    return m.group(1) if m else ''


def should_skip_invoice_line(line: str) -> bool:
    low = line.lower()
    bad = (
        'print date', 'moroccanoil israel ltd', 'moroccanoil', 'moshe levi', 'rishon lezion',
        'vat id', 'document date', 'po ref', 'shipping to', 'bill to', 'ctc person', 'email', 'tel',
        'unit price', '# item no.', 'quantity country', 'discount discount', 'page ',
        'printed by sap business one', 'balance due', 'total due', 'shipment method', 'payment terms',
        'inco terms', 'i declare that the above information', 'credit / paid', 'total lines before',
        'total after discount', 'document discount', 'total lines discount', 'vat 0', 'total $',
    )
    return low.startswith(bad)


def try_parse_invoice_single_line(line: str, shipment_key: str, file_name: str, original_name: str) -> Optional[Dict[str, Any]]:
    m = INVOICE_ROW_SINGLE_RE.match(line)
    if m:
        item_no = normalize_sku(m.group(1))
        item_index = int(m.group(2))
        quantity = parse_decimal(m.group(3))
        discount_pct = parse_decimal(m.group(4))
        description = normalize_space(m.group(5))
        unit_before = parse_decimal(m.group(6))
        total_before = parse_decimal(m.group(7))
        unit_after = parse_decimal(m.group(8))
        total_after = parse_decimal(m.group(9))
        commercial_discount = parse_decimal(m.group(10))
        country = normalize_space(m.group(11))
        return {
            'itemIndex': item_index,
            'itemNo': item_no,
            'description': description,
            'quantity': dec_to_num(quantity),
            'unitPriceBeforeDiscount': dec_to_num(unit_before),
            'totalBeforeDiscount': dec_to_num(total_before),
            'discountPercentage': dec_to_num(discount_pct),
            'unitPriceAfterDiscount': dec_to_num(unit_after),
            'totalPriceAfterDiscount': dec_to_num(total_after),
            'commercialDiscount': dec_to_num(commercial_discount),
            'countryOfOrigin': country,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
        }

    m2 = INVOICE_ROW_SINGLE_RE_NOT_APPLICABLE.match(line)
    if m2:
        item_no = normalize_sku(m2.group(1))
        item_index = int(m2.group(2))
        quantity = parse_decimal(m2.group(3))
        discount_pct = parse_decimal(m2.group(4))
        description = normalize_space(m2.group(5))
        country = normalize_space(m2.group(6))
        unit_before = parse_decimal(m2.group(7))
        total_before = parse_decimal(m2.group(8))
        unit_after = parse_decimal(m2.group(9))
        total_after = parse_decimal(m2.group(10))
        commercial_discount = parse_decimal(m2.group(11))
        return {
            'itemIndex': item_index,
            'itemNo': item_no,
            'description': description,
            'quantity': dec_to_num(quantity),
            'unitPriceBeforeDiscount': dec_to_num(unit_before),
            'totalBeforeDiscount': dec_to_num(total_before),
            'discountPercentage': dec_to_num(discount_pct),
            'unitPriceAfterDiscount': dec_to_num(unit_after),
            'totalPriceAfterDiscount': dec_to_num(total_after),
            'commercialDiscount': dec_to_num(commercial_discount),
            'countryOfOrigin': country,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
        }

    return None


def is_sku_continuation_token(value: str) -> bool:
    token = normalize_space(value)
    if not re.fullmatch(r'[A-Z0-9-]{1,8}', token):
        return False
    if token in {'FOC', 'DG', 'NDG'}:
        return False
    # A wrapped description can start with "75ML"; that is not a SKU suffix.
    if re.fullmatch(r'\d{2,4}(?:ML|MM|G|KG|OZ)', token):
        return False
    return True


def split_fno_continuation(
    lines: Sequence[str],
    start_index: int,
    initial_country: str,
) -> Tuple[str, str, str, Optional[str], int]:
    suffix = ''
    description_parts: List[str] = []
    country = normalize_space(initial_country)
    barcode: Optional[str] = None
    index = start_index
    first_content_line = True

    while index < len(lines):
        line = normalize_space(lines[index])
        if not line:
            index += 1
            continue
        if FNO_INVOICE_ROW_RE.match(line) or FNO_PACKING_ROW_RE.match(line):
            break
        if FNO_PACKAGE_HEADER_RE.match(line):
            break
        if should_skip_invoice_line(line) or should_skip_packing_line(line):
            break
        if line.lower().startswith(
            (
                'shipment method:',
                'total tax',
                'total line',
                'document discount',
                'total after',
                'vat ',
                'total ',
                '[for customs',
            )
        ):
            break

        tokens = line.split()
        if first_content_line and tokens and is_sku_continuation_token(tokens[0]):
            suffix = tokens.pop(0)
        first_content_line = False

        remaining: List[str] = []
        for token in tokens:
            digits = re.sub(r'\D', '', token)
            if 8 <= len(digits) <= 14 and digits == token:
                barcode = digits
            else:
                remaining.append(token)

        if country.lower() == 'united' and remaining and remaining[-1].lower() == 'states':
            country = 'United States'
            remaining.pop()
        elif country.lower() == 'not' and remaining and remaining[-1].lower() == 'applicable':
            country = 'Not Applicable'
            remaining.pop()

        if remaining:
            description_parts.append(' '.join(remaining))
        index += 1

    return suffix, normalize_space(' '.join(description_parts)), country, barcode, index


def parse_fno_invoice_rows(
    lines: Sequence[str],
    shipment_key: str,
    file_name: str,
    original_name: str,
    invoice_no: str,
    warnings: List[str],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    index = 0

    while index < len(lines):
        match = FNO_INVOICE_ROW_RE.match(normalize_space(lines[index]))
        if not match:
            index += 1
            continue

        item_index = int(match.group(1))
        is_foc = bool(match.group(2))
        item_no = normalize_sku(match.group(3))
        description = normalize_space(match.group(4))
        quantity = parse_decimal(match.group(5))
        unit_before = parse_decimal(match.group(6))
        total_before = parse_decimal(match.group(7))
        discount_pct = parse_decimal(match.group(8))
        unit_after = parse_decimal(match.group(9))
        total_after = parse_decimal(match.group(10))
        country = normalize_space(match.group(11))

        suffix, continuation_description, country, _, next_index = split_fno_continuation(
            lines, index + 1, country
        )
        if suffix and (len(suffix) == 1 or not item_no.endswith(suffix)):
            item_no = normalize_sku(item_no + suffix)
        if continuation_description:
            description = normalize_space(f'{description} {continuation_description}')

        if item_no == 'DGR':
            warnings.append(
                f'[invoice] Пропущена служебная строка DGR в {file_name}, строка {item_index}'
            )
            index = max(next_index, index + 1)
            continue

        commercial_discount = None
        if total_before is not None and total_after is not None:
            commercial_discount = total_before - total_after

        rows.append({
            'itemIndex': item_index,
            'itemNo': item_no,
            'description': description,
            'quantity': dec_to_num(quantity),
            'unitPriceBeforeDiscount': dec_to_num(unit_before),
            'totalBeforeDiscount': dec_to_num(total_before),
            'discountPercentage': dec_to_num(discount_pct),
            'unitPriceAfterDiscount': dec_to_num(unit_after),
            'totalPriceAfterDiscount': dec_to_num(total_after),
            'commercialDiscount': dec_to_num(commercial_discount),
            'countryOfOrigin': country,
            '__isFoc': is_foc,
            '__sourceLayout': 'fno-2026',
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
            '__invoiceNo': invoice_no,
        })
        index = max(next_index, index + 1)

    seen_line_numbers = set()
    for row_order, row in enumerate(rows, 1):
        line_number = row.get('itemIndex')
        row['__rowOrder'] = row_order
        row['__isComponent'] = line_number in seen_line_numbers
        seen_line_numbers.add(line_number)

    return rows


def parse_invoice_pdf(entry: Dict[str, Any], shipment_key: str, warnings: List[str]) -> Tuple[str, List[Dict[str, Any]]]:
    pdf_path = entry['path']
    file_name = os.path.basename(pdf_path)
    original_name = entry.get('original_name') or file_name
    lines = extract_pdf_lines(pdf_path)
    invoice_no = extract_invoice_no(lines, file_name)

    is_fno_layout = any(
        '# item no. description unit' in normalize_space(line).lower()
        and 'total price' in normalize_space(line).lower()
        for line in lines
    )
    if is_fno_layout:
        rows = parse_fno_invoice_rows(
            lines,
            shipment_key,
            file_name,
            original_name,
            invoice_no,
            warnings,
        )
        if not rows:
            warnings.append(f'[invoice] Не удалось распарсить строки invoice: {file_name}')
        if not invoice_no:
            warnings.append(f'[invoice] Не найден номер invoice в {file_name}')
        return invoice_no, rows

    rows: List[Dict[str, Any]] = []

    i = 0
    while i < len(lines) - 1:
        desc_line = lines[i]
        if should_skip_invoice_line(desc_line):
            i += 1
            continue
        desc_match = re.match(r'^(\d{1,3})\s+(.+)$', desc_line)
        if not desc_match:
            i += 1
            continue

        item_index = int(desc_match.group(1))
        description = normalize_space(desc_match.group(2))
        data_line = lines[i + 1]
        if should_skip_invoice_line(data_line):
            i += 1
            continue

        extra_sku_suffix = ''
        if i + 2 < len(lines) and re.fullmatch(r'[A-Z]{1,3}', lines[i + 2]):
            extra_sku_suffix = lines[i + 2]

        m = INVOICE_DATA_RE.match(data_line)
        if not m:
            i += 1
            continue

        item_no = normalize_sku(m.group(1) + extra_sku_suffix)
        desc_suffix = normalize_space(m.group(2))
        quantity = parse_decimal(m.group(3))
        unit_before = parse_decimal(m.group(4))
        total_before = parse_decimal(m.group(5))
        discount_pct = parse_decimal(m.group(6))
        unit_after = parse_decimal(m.group(7))
        total_after = parse_decimal(m.group(8))
        commercial_discount = parse_decimal(m.group(9))
        country = normalize_space(m.group(10))

        applicable_peek = i + 2 + (1 if extra_sku_suffix else 0)
        if country.lower() == 'not' and applicable_peek < len(lines) and lines[applicable_peek].strip().lower() == 'applicable':
            country = 'Not Applicable'

        if desc_suffix and len(desc_suffix) <= 20 and not re.search(r'\d', desc_suffix):
            description = f'{description} {desc_suffix}'.strip()

        rows.append({
            'itemIndex': item_index,
            'itemNo': item_no,
            'description': description,
            'quantity': dec_to_num(quantity),
            'unitPriceBeforeDiscount': dec_to_num(unit_before),
            'totalBeforeDiscount': dec_to_num(total_before),
            'discountPercentage': dec_to_num(discount_pct),
            'unitPriceAfterDiscount': dec_to_num(unit_after),
            'totalPriceAfterDiscount': dec_to_num(total_after),
            'commercialDiscount': dec_to_num(commercial_discount),
            'countryOfOrigin': country,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
            '__invoiceNo': invoice_no,
        })

        i += 2
        if extra_sku_suffix:
            i += 1
        if country == 'Not Applicable':
            i += 1

    if not rows:
        for line in lines:
            if should_skip_invoice_line(line):
                continue
            parsed = try_parse_invoice_single_line(line, shipment_key, file_name, original_name)
            if parsed:
                parsed['__invoiceNo'] = invoice_no
                rows.append(parsed)

    if not rows:
        warnings.append(f'[invoice] Не удалось распарсить строки invoice: {file_name}')
    if not invoice_no:
        warnings.append(f'[invoice] Не найден номер invoice в {file_name}')

    return invoice_no, rows


def clean_packing_description(text: str) -> str:
    text = re.sub(r'[█▐▌]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip(' -|')


def should_skip_packing_line(line: str) -> bool:
    low = line.lower()
    bad = (
        'packages content', 'moroccanoil israel ltd', 'moroccanoil il', 'moshe levi', 'rishon lezion',
        'customer name:', 'ship to:', 'customer ref. no', 'based on sales orders', 'package number',
        'item code item description', 'item code item description barcode weight qty total box',
        'item code item description barcode qty total box', 'subtotals:', 'grand total', 'net weight:',
        'signature', 'print date', 'page ', 'shipment:', 'via giorgio perlasca', 'martinengo', 'italy',
        '24057', 'i declare that the above informaton', 'i declare that the above information',
    )
    return low.startswith(bad)


def looks_like_noise(line: str) -> bool:
    stripped = normalize_space(line)
    if not stripped:
        return True
    return bool(re.fullmatch(r'[█▐▌\s]+', stripped))


def is_barcode_line(line: str) -> bool:
    return bool(re.fullmatch(r'\d{8,14}', normalize_space(line)))


def is_package_header_line(line: str) -> Optional[str]:
    m_old = PACKAGE_HEADER_RE_OLD.match(line)
    if m_old:
        return m_old.group(1)
    m_new = PACKAGE_HEADER_RE_NEW.match(line)
    if m_new:
        return m_new.group(1)
    return None


def is_machine_readable_noise(line: str) -> bool:
    text = normalize_space(line)
    if not text:
        return True
    return text.startswith('%') or bool(re.fullmatch(r'[^A-Za-zА-ЯЁа-яё0-9]+', text))


def parse_fno_packing_rows(
    lines: Sequence[str],
    shipment_key: str,
    file_name: str,
    original_name: str,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    current_package: Optional[str] = None
    current_sscc: Optional[str] = None
    current_gross_weight: Optional[float] = None
    current_net_weight: Optional[float] = None
    index = 0

    while index < len(lines):
        line = normalize_space(lines[index])
        header = FNO_PACKAGE_HEADER_RE.match(line)
        if header:
            current_package = header.group(1)
            current_sscc = header.group(2)
            current_gross_weight = dec_to_num(parse_decimal(header.group(8)))
            current_net_weight = dec_to_num(parse_decimal(header.group(9)))
            index += 1
            continue

        item = FNO_PACKING_ROW_RE.match(line)
        if not item:
            index += 1
            continue

        item_no = normalize_sku(item.group(1))
        description = extract_description_prefix(item.group(2))
        barcode: Optional[str] = item.group(3)
        quantity = dec_to_num(parse_decimal(item.group(4)))
        weight = dec_to_num(parse_decimal(item.group(6)))
        boxes = dec_to_num(parse_decimal(item.group(7)))
        suffix = ''
        description_parts: List[str] = []
        nested_in_cb: Optional[str] = None
        next_index = index + 1
        first_content_line = True

        while next_index < len(lines):
            continuation = normalize_space(lines[next_index])
            if FNO_PACKAGE_HEADER_RE.match(continuation):
                break
            if FNO_PACKING_ROW_RE.match(continuation):
                break
            if should_skip_packing_line(continuation):
                break
            if looks_like_noise(continuation) or is_machine_readable_noise(continuation):
                next_index += 1
                continue
            if continuation.lower().startswith('carton boxes:'):
                break

            nested_match = re.match(r'^In\s+(\d+)\b\s*(.*)$', continuation, re.I)
            if nested_match:
                nested_in_cb = nested_match.group(1)
                continuation = normalize_space(nested_match.group(2))
                if not continuation:
                    next_index += 1
                    continue

            tokens = continuation.split()
            if first_content_line and tokens and is_sku_continuation_token(tokens[0]):
                suffix = tokens.pop(0)
            first_content_line = False

            remaining: List[str] = []
            for token in tokens:
                digits = re.sub(r'\D', '', token)
                if 8 <= len(digits) <= 14 and digits == token:
                    barcode = digits
                else:
                    remaining.append(token)
            if remaining:
                description_parts.append(' '.join(remaining))
            next_index += 1

        if suffix and (len(suffix) == 1 or not item_no.endswith(suffix)):
            item_no = normalize_sku(item_no + suffix)
        if description_parts:
            description = normalize_space(f"{description} {' '.join(description_parts)}")

        rows.append({
            'itemNo': item_no,
            'descriptionFromPacking': description,
            'quantity': quantity,
            'weight': weight,
            'boxes': boxes,
            'barcode': barcode,
            'pallet': current_package,
            'sscc': current_sscc,
            'packageGrossWeight': current_gross_weight,
            'packageNetWeight': current_net_weight,
            'nestedInCb': nested_in_cb,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
        })
        index = max(next_index, index + 1)

    return rows


def is_new_packing_item_line(line: str) -> Optional[re.Match[str]]:
    return PACKING_ROW_RE_NEW.match(line)


def is_old_packing_item_line(line: str) -> Optional[re.Match[str]]:
    return PACKING_ROW_RE_OLD.match(line)


def extract_description_prefix(raw_desc: str) -> str:
    part = re.split(r'[█▐▌]+', raw_desc)[0]
    part = clean_packing_description(part)
    return part


def parse_packing_pdf(entry: Dict[str, Any], shipment_key: str, warnings: List[str]) -> List[Dict[str, Any]]:
    pdf_path = entry['path']
    file_name = os.path.basename(pdf_path)
    original_name = entry.get('original_name') or file_name
    lines = extract_pdf_lines(pdf_path)

    is_fno_layout = any(
        'item code item description barcode qty piece weight total box'
        in normalize_space(line).lower()
        for line in lines
    )
    if is_fno_layout:
        rows = parse_fno_packing_rows(lines, shipment_key, file_name, original_name)
        if not rows:
            warnings.append(f'[packing] Не удалось распарсить строки packing: {file_name}')
        return rows

    rows: List[Dict[str, Any]] = []
    current_pallet: Optional[str] = None
    pending_in_cb: Optional[str] = None
    is_new_layout = any('barcode weight qty total box' in normalize_space(line).lower() for line in lines)

    i = 0
    while i < len(lines):
        line = normalize_space(lines[i])

        header_pallet = is_package_header_line(line)
        if header_pallet:
            current_pallet = header_pallet
            pending_in_cb = None
            i += 1
            continue

        cb_match = IN_CB_RE.search(line)
        if cb_match:
            pending_in_cb = cb_match.group(1)
            i += 1
            continue

        if should_skip_packing_line(line) or looks_like_noise(line):
            i += 1
            continue

        new_match = is_new_packing_item_line(line)
        if is_new_layout and new_match:
            item_no = normalize_sku(new_match.group(1))
            first_desc = extract_description_prefix(new_match.group(2))
            weight = parse_decimal(new_match.group(3))
            quantity = parse_decimal(new_match.group(4))
            boxes = parse_decimal(new_match.group(5))
            desc_parts: List[str] = []
            if first_desc:
                desc_parts.append(first_desc)

            j = i + 1
            while j < len(lines):
                next_line = normalize_space(lines[j])
                if is_package_header_line(next_line):
                    break
                if should_skip_packing_line(next_line):
                    break
                if is_new_packing_item_line(next_line) or is_old_packing_item_line(next_line):
                    break
                if is_barcode_line(next_line):
                    j += 1
                    continue
                if looks_like_noise(next_line):
                    j += 1
                    continue
                cleaned = clean_packing_description(re.split(r'[█▐▌]+', next_line)[0])
                if cleaned and not re.fullmatch(r'[\d.,]+', cleaned):
                    desc_parts.append(cleaned)
                j += 1

            rows.append({
                'itemNo': item_no,
                'descriptionFromPacking': normalize_space(' '.join(desc_parts)),
                'quantity': dec_to_num(quantity),
                'weight': dec_to_num(weight),
                'boxes': dec_to_num(boxes),
                'pallet': current_pallet,
                'nestedInCb': pending_in_cb,
                '__sourceFileName': file_name,
                '__sourceOriginalName': original_name,
                '__shipmentKey': shipment_key,
            })
            pending_in_cb = None
            i = j
            continue

        old_match = is_old_packing_item_line(line)
        if old_match:
            rows.append({
                'itemNo': normalize_sku(old_match.group(1)),
                'descriptionFromPacking': clean_packing_description(old_match.group(2)),
                'quantity': dec_to_num(parse_decimal(old_match.group(3))),
                'weight': dec_to_num(parse_decimal(old_match.group(4))),
                'boxes': dec_to_num(parse_decimal(old_match.group(5))),
                'pallet': current_pallet,
                'nestedInCb': pending_in_cb,
                '__sourceFileName': file_name,
                '__sourceOriginalName': original_name,
                '__shipmentKey': shipment_key,
            })
            pending_in_cb = None
            i += 1
            continue

        i += 1

    if not rows:
        warnings.append(f'[packing] Не удалось распарсить строки packing: {file_name}')

    return rows


def sum_numeric(rows: Sequence[Dict[str, Any]], key: str) -> Decimal:
    total = Decimal('0')
    for row in rows:
        value = parse_decimal(row.get(key))
        if value is not None:
            total += value
    return total


def sum_invoice_product_quantity(rows: Sequence[Dict[str, Any]]) -> Decimal:
    return sum_numeric(
        [row for row in rows if not row.get('__isComponent')],
        'quantity',
    )


def has_document_count_mismatch(invoice_count: int, packing_count: int) -> bool:
    # One shipment may contain multiple invoices, but they all share one packing.
    return invoice_count < 1 or packing_count != 1


def format_excel_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    text = normalize_space(value)
    return text or None


def parse_batch_xlsx(entry: Dict[str, Any], shipment_key: str, warnings: List[str]) -> List[Dict[str, Any]]:
    xlsx_path = entry['path']
    file_name = os.path.basename(xlsx_path)
    original_name = entry.get('original_name') or file_name

    try:
        workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        warnings.append(f'[batch] Не удалось открыть batch файл {file_name}: {e}')
        return []

    worksheet = workbook[workbook.sheetnames[0]]
    all_rows = list(worksheet.iter_rows(values_only=True))
    header_index: Optional[int] = None
    header_map: Dict[str, int] = {}
    layout = ''

    for index, candidate in enumerate(all_rows):
        candidate_map = {
            normalize_header(header): column
            for column, header in enumerate(candidate)
            if header is not None
        }
        if {'sapitemcode', 'batchnum'}.issubset(candidate_map):
            header_index = index
            header_map = candidate_map
            layout = 'legacy'
            break
        if {'sku', 'batchno', 'qty'}.issubset(candidate_map):
            header_index = index
            header_map = candidate_map
            layout = 'shipping-data'
            break

    if header_index is None:
        warnings.append(f'[batch] Не найдена строка заголовков в {file_name}')
        return []

    def get_value(row: Sequence[Any], *header_names: str) -> Any:
        for header_name in header_names:
            idx = header_map.get(normalize_header(header_name))
            if idx is not None and idx < len(row):
                return row[idx]
        return None

    result: List[Dict[str, Any]] = []
    current_pallet = ''
    current_item_no = ''
    current_description = ''
    current_barcode: Optional[str] = None

    for row in all_rows[header_index + 1:]:
        if layout == 'shipping-data':
            current_pallet = normalize_space(get_value(row, 'Pallet')) or current_pallet
            current_item_no = normalize_sku(get_value(row, 'SKU')) or current_item_no
            current_description = normalize_space(get_value(row, 'Prod Name')) or current_description
            barcode_raw = get_value(row, 'EAN')
            parsed_barcode = re.sub(r'\D', '', str(barcode_raw or '')) or None
            current_barcode = parsed_barcode or current_barcode
            kit_component = normalize_space(get_value(row, 'Kit component'))
            is_kit_component = bool(
                kit_component and kit_component.upper() != 'N/A'
            )
            batch_no = normalize_space(get_value(row, 'Batch No'))
            if not current_item_no or not batch_no:
                continue
            quantity = dec_to_num(parse_decimal(get_value(row, 'Qty')))
            result.append({
                'itemNo': current_item_no,
                'wmsItemCode': '',
                'itemFrgnName': current_description,
                'quantity': quantity,
                'quantityUnit': 'pieces' if is_kit_component else 'boxes',
                'boxes': None if is_kit_component else quantity,
                'batchNo': batch_no,
                'kitBatchNo': normalize_space(get_value(row, 'Kit Batch No')) or None,
                'kitComponentDescription': kit_component if is_kit_component else None,
                'prodDate': format_excel_date(get_value(row, 'Prod. date')),
                'expDate': format_excel_date(get_value(row, 'Exp. date')),
                'barcode': current_barcode,
                'pallet': current_pallet or None,
                '__sourceFileName': file_name,
                '__sourceOriginalName': original_name,
                '__shipmentKey': shipment_key,
            })
            continue

        sap_item_code = get_value(row, 'SAP ItemCode', 'sapitemcode')
        item_no = normalize_sku(sap_item_code)
        if not item_no:
            continue
        barcode_raw = get_value(row, 'Barcode', 'barcode')
        barcode = re.sub(r'\D', '', str(barcode_raw or '')) or None
        result.append({
            'itemNo': item_no,
            'wmsItemCode': normalize_space(get_value(row, 'WMS ItemCode', 'wmsitemcode')),
            'itemFrgnName': normalize_space(get_value(row, 'ItemFrgnName', 'itemfrgnname')),
            'quantity': dec_to_num(parse_decimal(get_value(row, 'Quantity', 'quantity'))),
            'quantityUnit': 'pieces',
            'batchNo': normalize_space(get_value(row, 'BatchNum', 'batchnum')),
            'prodDate': format_excel_date(get_value(row, 'BatchProdDate', 'batchproddate')),
            'expDate': format_excel_date(get_value(row, 'BatchExpDate', 'batchexpdate')),
            'barcode': barcode,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
        })

    if not result:
        warnings.append(f'[batch] Не удалось распарсить строки batch: {file_name}')

    return result


def convert_batch_box_quantities_to_pieces(
    batch_rows: List[Dict[str, Any]],
    packing_rows: Sequence[Dict[str, Any]],
    warnings: List[str],
) -> None:
    packing_groups: Dict[Tuple[str, str], Dict[str, float]] = {}
    for row in packing_rows:
        item_no = normalize_sku(row.get('itemNo'))
        quantity = dec_to_num(parse_decimal(row.get('quantity')))
        boxes_raw = row.get('boxes')
        boxes = (
            0.0
            if boxes_raw in (0, 0.0, '0', '0.0')
            else dec_to_num(parse_decimal(boxes_raw))
        )
        identifiers = {
            normalize_space(row.get('sscc')),
            normalize_space(row.get('nestedInCb')),
            normalize_space(row.get('pallet')),
        }
        identifiers.discard('')
        if not item_no or not identifiers or quantity is None or boxes is None:
            continue
        for identifier in identifiers:
            group = packing_groups.setdefault(
                (item_no, identifier),
                {'quantity': 0.0, 'boxes': 0.0},
            )
            group['quantity'] += float(quantity)
            group['boxes'] += float(boxes)

    batch_groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for row in batch_rows:
        if row.get('quantityUnit') != 'boxes':
            continue
        item_no = normalize_sku(row.get('itemNo'))
        pallet = normalize_space(row.get('pallet'))
        if item_no and pallet:
            batch_groups.setdefault((item_no, pallet), []).append(row)

    for key, rows in batch_groups.items():
        packing = packing_groups.get(key)
        batch_boxes = sum(
            dec_to_num(parse_decimal(row.get('quantity'))) or 0.0 for row in rows
        )
        if not packing or batch_boxes <= 0:
            warnings.append(
                f'[batch] Не удалось сопоставить коробки batch с packing для {key[0]}, паллета {key[1]}'
            )
            continue
        if abs(packing['quantity'] - batch_boxes) <= 0.01:
            allocated_boxes = 0.0
            for row_index, row in enumerate(rows):
                pieces = float(
                    dec_to_num(parse_decimal(row.get('quantity'))) or 0.0
                )
                if row_index == len(rows) - 1:
                    boxes = packing['boxes'] - allocated_boxes
                else:
                    boxes = round(
                        packing['boxes'] * pieces / batch_boxes,
                        6,
                    )
                    allocated_boxes += boxes
                row['boxes'] = float(boxes)
                row['quantity'] = pieces
                row['quantityUnit'] = 'pieces'
            continue
        if abs(packing['boxes'] - batch_boxes) > 0.01:
            warnings.append(
                f'[batch] Коробки batch ({batch_boxes:g}) != packing ({packing["boxes"]:g}) '
                f'для {key[0]}, паллета {key[1]}'
            )
            continue

        converted_total = 0.0
        for row_index, row in enumerate(rows):
            boxes = dec_to_num(parse_decimal(row.get('quantity'))) or 0.0
            row['boxes'] = boxes
            if row_index == len(rows) - 1:
                pieces = packing['quantity'] - converted_total
            else:
                pieces = round(packing['quantity'] * boxes / batch_boxes, 6)
                converted_total += pieces
            row['quantity'] = float(pieces)
            row['quantityUnit'] = 'pieces'


def is_matching_batch_file(file_name: str, shipment_key: str) -> bool:
    lower_name = file_name.lower()
    shipment_key_lower = shipment_key.lower()
    patterns = [
        rf'^batch-{re.escape(shipment_key_lower)}\.xlsx$',
        rf'^batch-{re.escape(shipment_key_lower)}-\d+\.xlsx$',
        rf'^{re.escape(shipment_key_lower)}__batch-{re.escape(shipment_key_lower)}\.xlsx$',
        rf'^{re.escape(shipment_key_lower)}__batch-{re.escape(shipment_key_lower)}-\d+\.xlsx$',
        rf'^moil-batch-{re.escape(shipment_key_lower)}-\d+\.xlsx$',
    ]
    return any(re.fullmatch(pattern, lower_name) for pattern in patterns)


def discover_file_entries(input_dir: str, shipment_key: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], Optional[str]]:
    manifest_path, manifest = load_manifest(input_dir)
    manifest_entries = build_manifest_entries(input_dir, manifest)

    invoice_entries: List[Dict[str, Any]] = []
    packing_entries: List[Dict[str, Any]] = []
    batch_entries: List[Dict[str, Any]] = []

    if manifest_entries:
        for entry in manifest_entries:
            role = infer_file_role(entry['saved_name'], entry)
            if role == 'inv':
                invoice_entries.append(entry)
            elif role == 'pac':
                packing_entries.append(entry)
            elif role == 'batch':
                batch_entries.append(entry)
        return sorted(invoice_entries, key=lambda x: x['path']), sorted(packing_entries, key=lambda x: x['path']), sorted(batch_entries, key=lambda x: x['path']), manifest_path

    for root, _, files in os.walk(input_dir):
        for file_name in files:
            lower_name = file_name.lower()
            full_path = os.path.join(root, file_name)
            base_entry = {
                'path': full_path,
                'saved_name': file_name,
                'original_name': file_name,
                'doc_type': None,
                'meta': {},
            }

            if lower_name.endswith('.pdf'):
                if shipment_key not in file_name:
                    continue
                if '__inv-' in lower_name or lower_name.startswith('inv-') or lower_name.startswith(f'moil-inv-{shipment_key.lower()}-') or re.search(rf'(^|__|/)inv-{re.escape(shipment_key.lower())}-', lower_name):
                    invoice_entries.append(base_entry)
                elif '__pac-' in lower_name or lower_name.startswith('pac-') or lower_name.startswith(f'moil-pac-{shipment_key.lower()}-') or re.search(rf'(^|__|/)pac-{re.escape(shipment_key.lower())}-', lower_name):
                    packing_entries.append(base_entry)
            elif lower_name.endswith('.xlsx'):
                if is_matching_batch_file(lower_name, shipment_key):
                    batch_entries.append(base_entry)

    return sorted(invoice_entries, key=lambda x: x['path']), sorted(packing_entries, key=lambda x: x['path']), sorted(batch_entries, key=lambda x: x['path']), manifest_path


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Parse MOIL invoice + packing bundle')
    parser.add_argument('--shipment-key', dest='shipment_key')
    parser.add_argument('--shipment-ref', dest='shipment_key_alias')
    parser.add_argument('--invoice-pdf', action='append', default=[])
    parser.add_argument('--packing-pdf', action='append', default=[])
    parser.add_argument('--invoice', nargs='*', default=[])
    parser.add_argument('--packing', nargs='*', default=[])
    parser.add_argument('--input-dir')
    parser.add_argument('--pretty', action='store_true')
    return parser.parse_args(argv)


def flatten_file_args(values: Sequence[Any]) -> List[str]:
    result: List[str] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, (list, tuple)):
            for inner in value:
                if inner:
                    result.append(str(inner))
        else:
            if value:
                result.append(str(value))
    return result


def make_manual_entries(paths: Sequence[str], doc_type: str) -> List[Dict[str, Any]]:
    result: List[Dict[str, Any]] = []
    for path in paths:
        result.append({
            'path': path,
            'saved_name': os.path.basename(path),
            'original_name': os.path.basename(path),
            'doc_type': doc_type,
            'meta': {},
        })
    return result


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    shipment_key = normalize_space(args.shipment_key or args.shipment_key_alias)
    if not shipment_key:
        sys.stderr.write('Нужно передать --shipment-key или --shipment-ref\n')
        return 2

    invoice_entries = make_manual_entries(flatten_file_args([args.invoice_pdf, args.invoice]), 'inv')
    packing_entries = make_manual_entries(flatten_file_args([args.packing_pdf, args.packing]), 'pac')
    batch_entries: List[Dict[str, Any]] = []
    manifest_path: Optional[str] = None

    if args.input_dir:
        auto_invoice_entries, auto_packing_entries, auto_batch_entries, manifest_path = discover_file_entries(args.input_dir, shipment_key)
        if not invoice_entries:
            invoice_entries = auto_invoice_entries
        if not packing_entries:
            packing_entries = auto_packing_entries
        batch_entries = auto_batch_entries

    if not invoice_entries:
        sys.stderr.write('Не найден ни один invoice PDF\n')
        return 2
    if not packing_entries:
        sys.stderr.write('Не найден ни один packing PDF\n')
        return 2

    missing = [entry['path'] for entry in [*invoice_entries, *packing_entries, *batch_entries] if not os.path.exists(entry['path'])]
    if missing:
        sys.stderr.write('Не найдены файлы:\n' + '\n'.join(missing) + '\n')
        return 2

    warnings: List[str] = []
    invoice_rows: List[Dict[str, Any]] = []
    invoice_numbers: List[str] = []
    for entry in invoice_entries:
        invoice_no, rows = parse_invoice_pdf(entry, shipment_key, warnings)
        if invoice_no and invoice_no not in invoice_numbers:
            invoice_numbers.append(invoice_no)
        invoice_rows.extend(rows)

    packing_rows: List[Dict[str, Any]] = []
    for entry in packing_entries:
        packing_rows.extend(parse_packing_pdf(entry, shipment_key, warnings))

    batch_rows: List[Dict[str, Any]] = []
    for entry in batch_entries:
        batch_rows.extend(parse_batch_xlsx(entry, shipment_key, warnings))
    convert_batch_box_quantities_to_pieces(batch_rows, packing_rows, warnings)

    if has_document_count_mismatch(len(invoice_entries), len(packing_entries)):
        warnings.append(
            f'[validate] Требуется один или несколько invoice и ровно один общий packing; '
            f'получено invoice={len(invoice_entries)}, packing={len(packing_entries)}'
        )
    if not invoice_rows:
        warnings.append(f'[validate] По поставке {shipment_key} не распознано ни одной строки invoice')
    if not packing_rows:
        warnings.append(f'[validate] По поставке {shipment_key} не распознано ни одной строки packing')

    invoice_total_qty = sum_invoice_product_quantity(invoice_rows)
    packing_total_qty = sum_numeric(packing_rows, 'quantity')
    if invoice_total_qty and packing_total_qty and invoice_total_qty != packing_total_qty:
        warnings.append(f'[validate] Сумма quantity invoice ({invoice_total_qty}) != quantity packing ({packing_total_qty})')

    payload = {
        'shipmentKey': shipment_key,
        'invoiceNo': '; '.join(invoice_numbers),
        'invoiceDocsCount': len(invoice_entries),
        'packingDocsCount': len(packing_entries),
        'batchDocsCount': len(batch_entries),
        'invoiceFiles': [entry['path'] for entry in invoice_entries],
        'packingFiles': [entry['path'] for entry in packing_entries],
        'batchFiles': [entry['path'] for entry in batch_entries],
        'manifestPath': manifest_path,
        'manifestUsed': bool(manifest_path),
        'invoiceRows': invoice_rows,
        'packingRows': packing_rows,
        'batchRows': batch_rows,
        'warnings': warnings,
    }

    if args.pretty:
        json.dump(payload, sys.stdout, ensure_ascii=False, indent=2)
    else:
        json.dump(payload, sys.stdout, ensure_ascii=False, separators=(',', ':'))
    sys.stdout.write('\n')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())

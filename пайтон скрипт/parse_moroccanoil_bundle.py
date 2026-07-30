#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pdfplumber
from openpyxl import load_workbook

try:
    import xlrd
except Exception:
    xlrd = None


INVOICE_NO_RE = re.compile(r'Invoice No\.?\s*[:#]?\s*([0-9A-Za-z\-/]+)', re.I)
PACKAGE_HEADER_RE_INLINE = re.compile(r'^\s*(\d+)\s+((?:PL|CB)\d+)\s+(?:PL|CB)\b', re.I)
PACKAGE_HEADER_RE_SPLIT_CODE = re.compile(r'^((?:PL|CB)\d{8,})$', re.I)
FNO_TABLE_PACKAGE_RE = re.compile(
    r'^\s*(\d+)\s+((?:PL|CB)\d+)'
    r'(?:\s+(PL|CB))?'
    r'\s+(\d[\d,]*(?:\.\d+)?)'
    r'\s+(\d[\d,]*(?:\.\d+)?)'
    r'\s+(\d[\d,]*(?:\.\d+)?)'
    r'\s+(\d[\d,]*(?:\.\d+)?)'
    r'\s+(\d[\d,]*(?:\.\d+)?)'
    r'\s+(\d[\d,]*(?:\.\d+)?)\s*$',
    re.I,
)
CARTON_BOX_RE = re.compile(r'^Carton Box\s+(\d+)\s+(CB\d+)\s+(\d+(?:\.\d+)?)$', re.I)
INVOICE_DATA_RE = re.compile(
    r'^([A-Z0-9-]+)'
    r'(?:\s+(.+?))?'
    r'\s+(\d[\d,]*)'
    r'\s+(\d[\d,]*\.\d+)'
    r'\s+\$\s+(\d[\d,]*\.\d+)'
    r'\s+\$\s+(\d+(?:\.\d+)?)%'
    r'\s+(\d[\d,]*\.\d+)'
    r'\s+\$\s+(\d[\d,]*\.\d+)'
    r'\s+\$\s+(\d[\d,]*\.\d+)'
    r'\s+\$\s+([A-Za-z]+)$'
)
INVOICE_ROW_SINGLE_RE = re.compile(
    r'^([A-Z0-9-]+)\s+(\d+)\s+(\d[\d,]*)\s+(\d+(?:\.\d+)?)%'
    r'\s+(.+?)\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+'
    r'(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+(.+)$'
)
INVOICE_ROW_SINGLE_RE_NO_INDEX = re.compile(
    r'^([A-Z0-9-]+)\s+(\d[\d,]*)\s+(\d+(?:\.\d+)?)%'
    r'\s+(.+?)\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+'
    r'(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+(.+)$'
)
INVOICE_ROW_SINGLE_RE_NO_INDEX_ALT = re.compile(
    r'^([A-Z0-9-]+)\s+(.+?)\s+(\d[\d,]*)\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+'
    r'(\d+(?:\.\d+)?)%\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+(\d[\d,]*\.\d+)\s+\$\s+(.+)$'
)
PACKING_ROW_RE = re.compile(r'^([A-Z][A-Z0-9-]*)\s+(.+?)\s+(\d[\d,]*\.\d+)\s+(\d[\d,]*\.\d+)\s+(\d[\d,]*\.?\d*)\s*$')
BARCODE_RE = re.compile(r'^\d{8,14}$')
ONE_LETTER_RE = re.compile(r'^[A-Z]{1,3}$')
NOISE_RE = re.compile(r'^[█▐▌\s]+$')
STRICT_INVOICE_HEADERS = ['no', 'barcode', 'vendorcode', 'hscode', 'dgndg', 'descriptionofgoods', 'quantity']
STRICT_PACKING_HEADERS = ['no', 'barcode', 'vendercode', 'dgndg', 'descriptionofgoods', 'lot', 'quantity', 'netweight']
MANIFEST_FILE_NAMES = ('manifest.json', 'shipment_manifest.json')


def normalize_space(value: Any) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_code(value: Any) -> str:
    return normalize_space(value).upper()


def normalize_compact_code(value: Any) -> str:
    return re.sub(r'\s+', '', str(value or '')).upper()


def normalize_header(value: Any) -> str:
    return re.sub(r'[^a-z0-9а-яё]+', '', str(value or '').lower())


def parse_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    text = normalize_space(value).replace('$', '').replace('%', '').replace(' ', '').replace(',', '')
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def dec_to_num(value: Optional[Decimal]) -> Optional[float]:
    if value is None:
        return None
    return float(value)


def gtin_text(value: Any) -> Optional[str]:
    if value is None or value == '':
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = re.sub(r'\D', '', str(value))
    return text or None


def format_excel_date(value: Any) -> Optional[str]:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    text = normalize_space(value)
    if not text:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%y', '%d/%m/%Y', '%m/%d/%y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return text


def load_manifest(input_dir: str) -> Tuple[Optional[str], Optional[Dict[str, Any]]]:
    for file_name in MANIFEST_FILE_NAMES:
        path = os.path.join(input_dir, file_name)
        if not os.path.exists(path):
            continue
        try:
            with open(path, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
            if isinstance(data, dict):
                return path, data
        except Exception:
            pass
    return None, None


def infer_file_role(file_name: str, meta: Optional[Dict[str, Any]] = None) -> str:
    meta = meta or {}
    doc_type = normalize_space(meta.get('doc_type') or meta.get('file_type')).lower()
    if doc_type in {'inv', 'invoice'}:
        return 'inv'
    if doc_type in {'pac', 'packing'}:
        return 'pac'
    if doc_type == 'batch':
        return 'batch'

    low = file_name.lower()
    if re.search(r'(^|[-_])inv([-. _]|$)', low) or 'invoice' in low or ' ci ' in f' {low} ':
        return 'inv'
    if re.search(r'(^|[-_])pac([-. _]|$)', low) or 'pack' in low or ' pl ' in f' {low} ':
        return 'pac'
    if 'batch' in low:
        return 'batch'
    return 'unknown'


def build_manifest_entries(input_dir: str, manifest: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not isinstance(manifest, dict):
        return []
    input_root = os.path.realpath(input_dir)
    result: List[Dict[str, Any]] = []
    for raw in manifest.get('files') or []:
        if not isinstance(raw, dict):
            continue
        saved_path = normalize_space(raw.get('saved_path'))
        saved_name = normalize_space(raw.get('saved_name'))
        candidate = (
            saved_path
            if saved_path and os.path.isabs(saved_path)
            else os.path.join(input_root, saved_path or saved_name)
            if saved_path or saved_name
            else ''
        )
        path = os.path.realpath(candidate) if candidate else ''
        try:
            is_inside_input = (
                bool(path)
                and os.path.commonpath([input_root, path]) == input_root
            )
        except ValueError:
            is_inside_input = False
        if not is_inside_input or not os.path.isfile(path):
            continue
        result.append({
            'path': path,
            'saved_name': saved_name or os.path.basename(path),
            'original_name': normalize_space(raw.get('original_name')) or None,
            'doc_type': normalize_space(raw.get('doc_type')) or None,
            'meta': raw,
        })
    return result


def extract_pdf_lines(pdf_path: str) -> List[str]:
    result: List[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=2) or ''
            for raw in text.splitlines():
                line = normalize_space(raw)
                if line:
                    result.append(line)
    return result


def extract_pdf_tables(pdf_path: str) -> List[List[List[Any]]]:
    result: List[List[List[Any]]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            result.extend(page.extract_tables() or [])
    return result


def extract_invoice_no(lines: Sequence[str], fallback_file_name: str) -> str:
    full_text = '\n'.join(lines)
    match = INVOICE_NO_RE.search(full_text)
    if match:
        return normalize_space(match.group(1))
    m = re.search(r'(\d{6,})', fallback_file_name)
    return m.group(1) if m else ''


def should_skip_invoice_line(line: str) -> bool:
    low = line.lower()
    bad = (
        'print date', 'moroccanoil', 'vat id', 'document date', 'po ref', 'shipping to', 'bill to',
        'ctc person', 'email', 'tel', 'unit price', '# item no.', 'quantity country', 'discount discount',
        'page ', 'printed by sap business one', 'balance due', 'total due', 'shipment method',
        'payment terms', 'inco terms', 'i declare that the above information', 'credit / paid',
        'total lines before', 'total after discount', 'document discount', 'total lines discount',
        'vat 0', 'total $',
    )
    return low.startswith(bad)


def try_parse_invoice_single_line(line: str, shipment_key: str, file_name: str, original_name: str) -> Optional[Dict[str, Any]]:
    m = INVOICE_ROW_SINGLE_RE.match(line)
    if not m:
        return None
    return {
        'itemIndex': int(m.group(2)),
        'itemNo': normalize_code(m.group(1)),
        'description': normalize_space(m.group(5)),
        'quantity': dec_to_num(parse_decimal(m.group(3))),
        'unitPriceBeforeDiscount': dec_to_num(parse_decimal(m.group(6))),
        'totalBeforeDiscount': dec_to_num(parse_decimal(m.group(7))),
        'discountPercentage': dec_to_num(parse_decimal(m.group(4))),
        'unitPriceAfterDiscount': dec_to_num(parse_decimal(m.group(8))),
        'totalPriceAfterDiscount': dec_to_num(parse_decimal(m.group(9))),
        'commercialDiscount': dec_to_num(parse_decimal(m.group(10))),
        'countryOfOrigin': normalize_space(m.group(11)),
        '__sourceFileName': file_name,
        '__sourceOriginalName': original_name,
        '__shipmentKey': shipment_key,
    }


def try_parse_invoice_single_line_no_index(
    line: str,
    shipment_key: str,
    file_name: str,
    original_name: str,
) -> Optional[Dict[str, Any]]:
    m = INVOICE_ROW_SINGLE_RE_NO_INDEX.match(line)
    if m:
        return {
            'itemIndex': None,
            'itemNo': normalize_code(m.group(1)),
            'description': normalize_space(m.group(4)),
            'quantity': dec_to_num(parse_decimal(m.group(2))),
            'unitPriceBeforeDiscount': dec_to_num(parse_decimal(m.group(5))),
            'totalBeforeDiscount': dec_to_num(parse_decimal(m.group(6))),
            'discountPercentage': dec_to_num(parse_decimal(m.group(3))),
            'unitPriceAfterDiscount': dec_to_num(parse_decimal(m.group(7))),
            'totalPriceAfterDiscount': dec_to_num(parse_decimal(m.group(8))),
            'commercialDiscount': dec_to_num(parse_decimal(m.group(9))),
            'countryOfOrigin': normalize_space(m.group(10)),
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
        }

    m = INVOICE_ROW_SINGLE_RE_NO_INDEX_ALT.match(line)
    if m:
        return {
            'itemIndex': None,
            'itemNo': normalize_code(m.group(1)),
            'description': normalize_space(m.group(2)),
            'quantity': dec_to_num(parse_decimal(m.group(3))),
            'unitPriceBeforeDiscount': dec_to_num(parse_decimal(m.group(4))),
            'totalBeforeDiscount': dec_to_num(parse_decimal(m.group(5))),
            'discountPercentage': dec_to_num(parse_decimal(m.group(6))),
            'unitPriceAfterDiscount': dec_to_num(parse_decimal(m.group(7))),
            'totalPriceAfterDiscount': dec_to_num(parse_decimal(m.group(8))),
            'commercialDiscount': dec_to_num(parse_decimal(m.group(9))),
            'countryOfOrigin': normalize_space(m.group(10)),
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
        }

    return None


def annotate_component_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    for idx, row in enumerate(rows):
        row['__rowOrder'] = idx + 1
        row['__isComponent'] = row.get('itemIndex') is None
    return rows


def find_table_header(
    table: Sequence[Sequence[Any]],
    required_headers: Sequence[str],
) -> Tuple[Optional[int], Dict[str, int]]:
    for row_index, row in enumerate(table):
        header_map: Dict[str, int] = {}
        for column_index, cell in enumerate(row):
            normalized = normalize_header(cell)
            if normalized:
                header_map[normalized] = column_index
        if all(header in header_map for header in required_headers):
            return row_index, header_map
    return None, {}


def parse_fno_invoice_tables(
    tables: Sequence[Sequence[Sequence[Any]]],
    shipment_key: str,
    file_name: str,
    original_name: str,
    invoice_no: str,
) -> List[Dict[str, Any]]:
    required_headers = (
        'itemno',
        'description',
        'unitquantity',
        'unitpricebeforediscount',
        'totalbeforediscount',
        'discount',
        'unitpriceafterdiscount',
        'totalprice',
        'producecountry',
    )
    result: List[Dict[str, Any]] = []

    for table in tables:
        header_index, header_map = find_table_header(table, required_headers)
        if header_index is None:
            continue

        for raw_row in table[header_index + 1:]:
            row = list(raw_row)

            def value(header: str) -> Any:
                column = header_map[header]
                return row[column] if column < len(row) else None

            index_text = normalize_space(row[0] if row else None)
            index_match = re.fullmatch(r'(\d+)(?:\s+FOC)?', index_text, re.I)
            if not index_match:
                continue

            item_no = normalize_compact_code(value('itemno'))
            quantity = parse_decimal(value('unitquantity'))
            if not item_no or quantity is None:
                continue

            total_before = parse_decimal(value('totalbeforediscount'))
            total_after = parse_decimal(value('totalprice'))
            discount = parse_decimal(value('discount'))
            commercial_discount = (
                total_before - total_after
                if total_before is not None and total_after is not None
                else None
            )
            is_foc = bool(
                re.search(r'\bFOC\b', index_text, re.I)
                or discount == Decimal('100')
            )

            result.append({
                'itemIndex': int(index_match.group(1)),
                'itemNo': item_no,
                'description': normalize_space(value('description')),
                'quantity': dec_to_num(quantity),
                'unitPriceBeforeDiscount': dec_to_num(
                    parse_decimal(value('unitpricebeforediscount'))
                ),
                'totalBeforeDiscount': dec_to_num(total_before),
                'discountPercentage': dec_to_num(discount),
                'unitPriceAfterDiscount': dec_to_num(
                    parse_decimal(value('unitpriceafterdiscount'))
                ),
                'totalPriceAfterDiscount': dec_to_num(total_after),
                'commercialDiscount': dec_to_num(commercial_discount),
                'countryOfOrigin': normalize_space(value('producecountry')),
                '__isFoc': is_foc,
                '__sourceLayout': 'fno-table-2026',
                '__sourceFileName': file_name,
                '__sourceOriginalName': original_name,
                '__shipmentKey': shipment_key,
                '__invoiceNo': invoice_no,
            })

    return annotate_component_rows(result)


def parse_invoice_pdf(entry: Dict[str, Any], shipment_key: str, warnings: List[str]) -> Tuple[str, List[Dict[str, Any]]]:
    pdf_path = entry['path']
    file_name = os.path.basename(pdf_path)
    original_name = entry.get('original_name') or file_name
    lines = extract_pdf_lines(pdf_path)
    invoice_no = extract_invoice_no(lines, file_name)

    try:
        table_rows = parse_fno_invoice_tables(
            extract_pdf_tables(pdf_path),
            shipment_key,
            file_name,
            original_name,
            invoice_no,
        )
    except Exception as error:
        table_rows = []
        warnings.append(
            f'[invoice] Не удалось прочитать табличный слой {file_name}: {error}'
        )
    if table_rows:
        return invoice_no, table_rows

    rows: List[Dict[str, Any]] = []

    def _row_from_data_match(
        m: re.Match,
        item_index: Optional[int],
        description: str,
        advance: int,
    ) -> Tuple[Dict[str, Any], int]:
        """Build a row dict from INVOICE_DATA_RE match; also handle extra SKU suffix and
        'Not Applicable' country split. Returns (row, new_advance)."""
        extra_sku = ''
        if advance < len(lines) and ONE_LETTER_RE.fullmatch(lines[advance]):
            extra_sku = lines[advance]
            advance += 1
        country = normalize_space(m.group(10))
        if country.lower() == 'not' and advance < len(lines) and lines[advance].strip().lower() == 'applicable':
            country = 'Not Applicable'
            advance += 1
        item_no = normalize_code(m.group(1) + extra_sku)
        desc_suffix = normalize_space(m.group(2) or '')
        if desc_suffix and len(desc_suffix) <= 30:
            description = f'{description} {desc_suffix}'.strip()
        row = {
            'itemIndex': item_index,
            'itemNo': item_no,
            'description': description,
            'quantity': dec_to_num(parse_decimal(m.group(3))),
            'unitPriceBeforeDiscount': dec_to_num(parse_decimal(m.group(4))),
            'totalBeforeDiscount': dec_to_num(parse_decimal(m.group(5))),
            'discountPercentage': dec_to_num(parse_decimal(m.group(6))),
            'unitPriceAfterDiscount': dec_to_num(parse_decimal(m.group(7))),
            'totalPriceAfterDiscount': dec_to_num(parse_decimal(m.group(8))),
            'commercialDiscount': dec_to_num(parse_decimal(m.group(9))),
            'countryOfOrigin': country,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
            '__invoiceNo': invoice_no,
        }
        return row, advance

    i = 0
    while i < len(lines):
        line = lines[i]

        if should_skip_invoice_line(line):
            i += 1
            continue

        # Old single-line formats (index as 2nd field, or no index with discount% before prices)
        parsed = try_parse_invoice_single_line(line, shipment_key, file_name, original_name)
        if not parsed:
            parsed = try_parse_invoice_single_line_no_index(line, shipment_key, file_name, original_name)
        if parsed:
            parsed['__invoiceNo'] = invoice_no
            rows.append(parsed)
            i += 1
            continue

        # Pattern A: "{index} {description}" line followed by INVOICE_DATA_RE data line
        desc_match = re.match(r'^(\d{1,3})\s+(.+)$', line)
        if desc_match and not INVOICE_DATA_RE.match(line):
            item_index = int(desc_match.group(1))
            description = normalize_space(desc_match.group(2))
            j = i + 1
            while j < len(lines) and (should_skip_invoice_line(lines[j]) or not lines[j].strip()):
                j += 1
            if j < len(lines):
                m = INVOICE_DATA_RE.match(lines[j])
                if m:
                    row, advance = _row_from_data_match(m, item_index, description, j + 1)
                    rows.append(row)
                    i = advance
                    continue
            i += 1
            continue

        # Pattern B: description text line followed immediately by INVOICE_DATA_RE data line
        j = i + 1
        while j < len(lines) and (should_skip_invoice_line(lines[j]) or not lines[j].strip()):
            j += 1
        if j < len(lines):
            m = INVOICE_DATA_RE.match(lines[j])
            if m:
                row, advance = _row_from_data_match(m, None, normalize_space(line), j + 1)
                rows.append(row)
                i = advance
                continue

        # Pattern C: standalone INVOICE_DATA_RE data line (no preceding description)
        m = INVOICE_DATA_RE.match(line)
        if m:
            row, advance = _row_from_data_match(m, None, '', i + 1)
            rows.append(row)
            i = advance
            continue

        i += 1

    if not rows:
        warnings.append(f'[invoice] Не удалось распарсить строки invoice: {file_name}')
    if not invoice_no:
        warnings.append(f'[invoice] Не найден номер invoice в {file_name}')

    return invoice_no, annotate_component_rows(rows)


def looks_like_noise(line: str) -> bool:
    return not line or NOISE_RE.fullmatch(line) is not None


def should_skip_packing_line(line: str) -> bool:
    low = line.lower()
    bad = (
        'packages content', 'moroccanoil il', 'customer name:', 'ship to:', 'customer ref. no',
        'based on sales orders', 'package number', 'item code item description', 'subtotals:',
        'page ', 'signature :', 'i declare that the above informaton', 'grand total', 'net weight:',
    )
    return low.startswith(bad)


def clean_packing_description(text: str) -> str:
    text = re.sub(r'[█▐▌]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip(' -|')


def extract_desc_prefix(text: str) -> str:
    return clean_packing_description(re.split(r'[█▐▌]+', text)[0])


def parse_fno_packing_tables(
    tables: Sequence[Sequence[Sequence[Any]]],
    shipment_key: str,
    file_name: str,
    original_name: str,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    current_package: Optional[str] = None
    current_package_number: Optional[int] = None
    current_package_type: Optional[str] = None
    current_gross_weight: Optional[float] = None
    current_net_weight: Optional[float] = None

    for table in tables:
        for raw_row in table:
            row = list(raw_row)
            if not row:
                continue

            first_cell = normalize_space(row[0])
            package_match = FNO_TABLE_PACKAGE_RE.fullmatch(first_cell)
            if package_match:
                current_package_number = int(package_match.group(1))
                current_package = normalize_compact_code(package_match.group(2))
                current_package_type = normalize_code(
                    package_match.group(3) or current_package[:2]
                )
                current_gross_weight = dec_to_num(
                    parse_decimal(package_match.group(8))
                )
                current_net_weight = dec_to_num(
                    parse_decimal(package_match.group(9))
                )
                continue

            if current_package is None or len(row) < 6:
                continue

            item_no = normalize_compact_code(row[0])
            if not re.fullmatch(r'[A-Z][A-Z0-9-]{3,}', item_no):
                continue

            quantity = parse_decimal(row[-4])
            weight = parse_decimal(row[-2])
            boxes = parse_decimal(row[-1])
            if quantity is None or weight is None or boxes is None:
                continue

            barcode: Optional[str] = None
            for cell in row[2:-4]:
                candidate = normalize_space(cell)
                if BARCODE_RE.fullmatch(candidate):
                    barcode = candidate
                    break

            rows.append({
                'itemNo': item_no,
                'descriptionFromPacking': normalize_space(
                    row[1] if len(row) > 1 else None
                ),
                'quantity': dec_to_num(quantity),
                'weight': dec_to_num(weight),
                'boxes': dec_to_num(boxes),
                'barcode': barcode,
                'pallet': current_package,
                'sscc': current_package,
                'packageNumber': current_package_number,
                'packageType': current_package_type,
                'packageGrossWeight': current_gross_weight,
                'packageNetWeight': current_net_weight,
                'nestedInCb': None,
                '__sourceLayout': 'fno-table-2026',
                '__sourceFileName': file_name,
                '__sourceOriginalName': original_name,
                '__shipmentKey': shipment_key,
            })

    return rows


def parse_packing_pdf(entry: Dict[str, Any], shipment_key: str, warnings: List[str]) -> List[Dict[str, Any]]:
    pdf_path = entry['path']
    file_name = os.path.basename(pdf_path)
    original_name = entry.get('original_name') or file_name

    try:
        table_rows = parse_fno_packing_tables(
            extract_pdf_tables(pdf_path),
            shipment_key,
            file_name,
            original_name,
        )
    except Exception as error:
        table_rows = []
        warnings.append(
            f'[packing] Не удалось прочитать табличный слой {file_name}: {error}'
        )
    if table_rows:
        return table_rows

    lines = extract_pdf_lines(pdf_path)

    rows: List[Dict[str, Any]] = []
    current_pallet: Optional[str] = None
    pending_cb_number: Optional[str] = None
    pending_cb_boxes: Optional[float] = None

    i = 0
    while i < len(lines):
        line = normalize_space(lines[i])
        if looks_like_noise(line) or should_skip_packing_line(line):
            i += 1
            continue

        m_header = PACKAGE_HEADER_RE_INLINE.match(line)
        if m_header:
            current_pallet = m_header.group(2)
            pending_cb_number = None
            pending_cb_boxes = None
            i += 1
            continue

        m_split_code = PACKAGE_HEADER_RE_SPLIT_CODE.match(line)
        if m_split_code and i + 1 < len(lines):
            current_pallet = m_split_code.group(1)
            pending_cb_number = None
            pending_cb_boxes = None
            i += 1
            continue

        m_carton = CARTON_BOX_RE.match(line)
        if m_carton:
            pending_cb_number = m_carton.group(1)
            try:
                pending_cb_boxes = float(m_carton.group(3))
            except Exception:
                pending_cb_boxes = None
            i += 1
            continue

        m_row = PACKING_ROW_RE.match(line)
        if not m_row:
            i += 1
            continue

        desc_parts: List[str] = []
        barcode: Optional[str] = None
        desc_start = extract_desc_prefix(m_row.group(2))
        if desc_start:
            desc_parts.append(desc_start)

        j = i + 1
        while j < len(lines):
            nxt = normalize_space(lines[j])
            if not nxt:
                j += 1
                continue
            if looks_like_noise(nxt):
                j += 1
                continue
            if PACKAGE_HEADER_RE_INLINE.match(nxt) or PACKAGE_HEADER_RE_SPLIT_CODE.match(nxt):
                break
            if should_skip_packing_line(nxt) or CARTON_BOX_RE.match(nxt) or PACKING_ROW_RE.match(nxt):
                break
            if BARCODE_RE.fullmatch(nxt):
                barcode = nxt
                j += 1
                continue
            cleaned = clean_packing_description(re.split(r'[█▐▌]+', nxt)[0])
            if cleaned and not re.fullmatch(r'[\d.,]+', cleaned):
                desc_parts.append(cleaned)
            j += 1

        boxes = dec_to_num(parse_decimal(m_row.group(5)))
        if boxes == 0 and pending_cb_boxes not in (None, 0):
            boxes = pending_cb_boxes

        rows.append({
            'itemNo': normalize_code(m_row.group(1)),
            'descriptionFromPacking': normalize_space(' '.join(desc_parts)),
            'quantity': dec_to_num(parse_decimal(m_row.group(4))),
            'weight': dec_to_num(parse_decimal(m_row.group(3))),
            'boxes': boxes,
            'barcode': barcode,
            'pallet': current_pallet,
            'nestedInCb': pending_cb_number,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
        })

        pending_cb_number = None
        pending_cb_boxes = None
        i = j

    if not rows:
        warnings.append(f'[packing] Не удалось распарсить строки packing: {file_name}')

    return rows


def resolve_sheet_names(all_sheet_names: Sequence[str], preferred_sheets: Sequence[str]) -> List[str]:
    if not preferred_sheets:
        return list(all_sheet_names)
    exact = [name for name in preferred_sheets if name in all_sheet_names]
    if exact:
        return exact
    preferred_norm = {normalize_header(name) for name in preferred_sheets}
    fuzzy = [name for name in all_sheet_names if normalize_header(name) in preferred_norm]
    if fuzzy:
        return fuzzy
    return list(all_sheet_names)


def iter_excel_rows(path: str, preferred_sheets: Sequence[str] = ()) -> Iterable[Tuple[str, List[Any]]]:
    lower = path.lower()
    if lower.endswith('.xlsx'):
        wb = load_workbook(path, data_only=True, read_only=True)
        ordered = resolve_sheet_names(list(wb.sheetnames), preferred_sheets)
        for sheet_name in ordered:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                yield sheet_name, list(row)
    elif lower.endswith('.xls'):
        if xlrd is None:
            raise RuntimeError('Для чтения .xls нужен пакет xlrd. Установи: /opt/moil-venv/bin/pip install xlrd')
        book = xlrd.open_workbook(path)
        ordered = resolve_sheet_names(book.sheet_names(), preferred_sheets)
        for sheet_name in ordered:
            sh = book.sheet_by_name(sheet_name)
            for rx in range(sh.nrows):
                yield sheet_name, sh.row_values(rx)
    else:
        raise RuntimeError(f'Неподдерживаемое расширение файла: {path}')


def find_header_row(rows: List[List[Any]], required_headers: Sequence[str]) -> Tuple[int, Dict[str, int]]:
    for idx, row in enumerate(rows):
        header_map: Dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            norm = normalize_header(cell)
            if norm:
                header_map[norm] = col_idx
        if all(h in header_map for h in required_headers):
            return idx, header_map
    raise RuntimeError(f'Не найдена строка заголовков. Ожидались колонки: {", ".join(required_headers)}')


def parse_invoice_xlsx_strict(entry: Dict[str, Any], shipment_key: str, warnings: List[str]) -> Tuple[str, List[Dict[str, Any]]]:
    path = entry['path']
    file_name = os.path.basename(path)
    original_name = entry.get('original_name') or file_name
    all_rows: List[List[Any]] = [row for _, row in iter_excel_rows(path)]
    try:
        header_idx, header_map = find_header_row(all_rows, STRICT_INVOICE_HEADERS)
    except Exception:
        warnings.append(f'[invoice] XLSX invoice {file_name} не похож на поддерживаемый strict-формат')
        return '', []

    rows: List[Dict[str, Any]] = []
    qty_idx = header_map['quantity']
    for row in all_rows[header_idx + 1:]:
        line_no = parse_decimal(row[header_map['no']] if header_map['no'] < len(row) else None)
        gtin = gtin_text(row[header_map['barcode']] if header_map['barcode'] < len(row) else None)
        vendor = normalize_code(row[header_map['vendorcode']] if header_map['vendorcode'] < len(row) else None)
        qty = dec_to_num(parse_decimal(row[qty_idx] if qty_idx < len(row) else None))
        if not gtin and not vendor and line_no is None:
            continue
        if not vendor:
            continue

        after_qty = row[qty_idx + 1: qty_idx + 6]
        numeric_after_qty = [dec_to_num(parse_decimal(v)) for v in after_qty if dec_to_num(parse_decimal(v)) is not None]
        unit_price = numeric_after_qty[0] if len(numeric_after_qty) >= 1 else None
        amount = numeric_after_qty[1] if len(numeric_after_qty) >= 2 else None

        rows.append({
            'itemIndex': int(line_no) if line_no is not None else None,
            'itemNo': vendor,
            'description': normalize_space(row[header_map['descriptionofgoods']] if header_map['descriptionofgoods'] < len(row) else None) or None,
            'quantity': qty,
            'unitPriceBeforeDiscount': unit_price,
            'totalBeforeDiscount': amount,
            'discountPercentage': 0,
            'unitPriceAfterDiscount': unit_price,
            'totalPriceAfterDiscount': amount,
            'commercialDiscount': 0,
            'countryOfOrigin': None,
            'gtin': gtin,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
            '__invoiceNo': '',
        })

    if not rows:
        warnings.append(f'[invoice] Не удалось распарсить строки invoice xlsx: {file_name}')
    return '', annotate_component_rows(rows)


def parse_packing_xlsx_strict(entry: Dict[str, Any], shipment_key: str, warnings: List[str]) -> List[Dict[str, Any]]:
    path = entry['path']
    file_name = os.path.basename(path)
    original_name = entry.get('original_name') or file_name
    all_rows: List[List[Any]] = [row for _, row in iter_excel_rows(path)]
    try:
        header_idx, header_map = find_header_row(all_rows, STRICT_PACKING_HEADERS)
    except Exception:
        warnings.append(f'[packing] XLSX packing {file_name} не похож на поддерживаемый strict-формат')
        return []

    rows: List[Dict[str, Any]] = []
    current_pallet: Optional[str] = None
    for row in all_rows[header_idx + 1:]:
        line_no = parse_decimal(row[header_map['no']] if header_map['no'] < len(row) else None)
        gtin = gtin_text(row[header_map['barcode']] if header_map['barcode'] < len(row) else None)
        vendor = normalize_code(row[header_map['vendercode']] if header_map['vendercode'] < len(row) else None)
        lot = normalize_space(row[header_map['lot']] if header_map['lot'] < len(row) else None)
        qty = dec_to_num(parse_decimal(row[header_map['quantity']] if header_map['quantity'] < len(row) else None))
        net_weight = dec_to_num(parse_decimal(row[header_map['netweight']] if header_map['netweight'] < len(row) else None))
        if not gtin and not vendor and line_no is None:
            continue
        if not vendor:
            continue

        gross_weight = dec_to_num(parse_decimal(row[9])) if len(row) > 9 else None
        carton_no = normalize_space(row[13]) or None if len(row) > 13 else None
        pallet_no = normalize_space(row[15]) or None if len(row) > 15 else None
        if pallet_no:
            current_pallet = pallet_no
        else:
            pallet_no = current_pallet

        rows.append({
            'itemNo': vendor,
            'descriptionFromPacking': normalize_space(row[header_map['descriptionofgoods']] if header_map['descriptionofgoods'] < len(row) else None) or None,
            'quantity': qty,
            'weight': net_weight,
            'grossWeight': gross_weight,
            'boxes': dec_to_num(parse_decimal(carton_no)),
            'barcode': gtin,
            'pallet': pallet_no,
            'lot': lot or None,
            'nestedInCb': None,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
        })

    if not rows:
        warnings.append(f'[packing] Не удалось распарсить строки packing xlsx: {file_name}')
    return rows


def parse_batch_xlsx(entry: Dict[str, Any], shipment_key: str, warnings: List[str]) -> List[Dict[str, Any]]:
    xlsx_path = entry['path']
    file_name = os.path.basename(xlsx_path)
    original_name = entry.get('original_name') or file_name

    try:
        workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        warnings.append(f'[batch] Не удалось открыть batch файл {file_name}: {e}')
        return []

    worksheet = workbook[workbook.sheetnames[0]]
    all_rows = list(worksheet.iter_rows(values_only=True))
    if not all_rows:
        warnings.append(f'[batch] Пустой batch файл: {file_name}')
        return []

    header_index: Optional[int] = None
    header_map: Dict[str, int] = {}
    layout = ''
    for index, candidate in enumerate(all_rows):
        candidate_map = {
            normalize_header(header): column
            for column, header in enumerate(candidate)
            if header is not None
        }
        if {'containerid1', 'itemid1', 'qty'}.issubset(candidate_map):
            header_index = index
            header_map = candidate_map
            layout = 'fno-technical'
            break
        if {'sapitemcode', 'batchnum'}.issubset(candidate_map):
            header_index = index
            header_map = candidate_map
            layout = 'legacy'
            break

    if header_index is None:
        warnings.append(f'[batch] Не найдена строка заголовков в {file_name}')
        return []

    def get_value(row: Sequence[Any], *header_names: str) -> Any:
        for header_name in header_names:
            idx = header_map.get(normalize_header(header_name))
            if idx is not None and idx < len(row):
                return row[idx]
        return None

    def optional_text(value: Any) -> Optional[str]:
        text = normalize_space(value)
        if not text or text.upper() in {'N/A', 'NA', 'NONE', '-'}:
            return None
        return text

    result: List[Dict[str, Any]] = []
    for row in all_rows[header_index + 1:]:
        if layout == 'fno-technical':
            item_no = normalize_compact_code(get_value(row, 'ItemId1'))
            quantity = dec_to_num(parse_decimal(get_value(row, 'Qty')))
            if not item_no or quantity is None:
                continue

            kit_component = optional_text(get_value(row, 'KitItemName'))
            is_kit_component = bool(kit_component)
            result.append({
                'itemNo': item_no,
                'wmsItemCode': '',
                'itemFrgnName': normalize_space(get_value(row, 'ItemName2')),
                'quantity': quantity,
                'quantityUnit': 'pieces' if is_kit_component else 'boxes',
                'boxes': None if is_kit_component else quantity,
                'batchNo': optional_text(get_value(row, 'InventBatchId')),
                'kitBatchNo': optional_text(get_value(row, 'KitInventBatchId')),
                'kitComponentDescription': kit_component,
                'prodDate': format_excel_date(get_value(row, 'ProdDate')),
                'expDate': format_excel_date(get_value(row, 'ExpDate')),
                'barcode': gtin_text(get_value(row, 'ItemBarCode3')),
                'pallet': optional_text(get_value(row, 'ContainerId1')),
                '__sourceLayout': 'fno-technical-2026',
                '__sourceFileName': file_name,
                '__sourceOriginalName': original_name,
                '__shipmentKey': shipment_key,
            })
            continue

        sap_item_code = get_value(row, 'SAP ItemCode', 'sapitemcode')
        item_no = normalize_code(sap_item_code)
        if not item_no:
            continue
        result.append({
            'itemNo': item_no,
            'wmsItemCode': normalize_space(get_value(row, 'WMS ItemCode', 'wmsitemcode')),
            'itemFrgnName': normalize_space(get_value(row, 'ItemFrgnName', 'itemfrgnname')),
            'quantity': dec_to_num(parse_decimal(get_value(row, 'Quantity', 'quantity'))),
            'quantityUnit': 'pieces',
            'batchNo': optional_text(get_value(row, 'BatchNum', 'batchnum')),
            'prodDate': format_excel_date(get_value(row, 'BatchProdDate', 'batchproddate')),
            'expDate': format_excel_date(get_value(row, 'BatchExpDate', 'batchexpdate')),
            'barcode': gtin_text(get_value(row, 'Barcode', 'barcode')),
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
        })

    if not result:
        warnings.append(f'[batch] Не удалось распарсить строки batch: {file_name}')
    return result


def convert_batch_box_quantities_to_pieces(
    batch_rows: List[Dict[str, Any]],
    packing_rows: Sequence[Dict[str, Any]],
    warnings: List[str],
) -> None:
    packing_groups: Dict[Tuple[str, str], Dict[str, float]] = {}
    for row in packing_rows:
        item_no = normalize_compact_code(row.get('itemNo'))
        quantity = dec_to_num(parse_decimal(row.get('quantity')))
        boxes_value = row.get('boxes')
        boxes = (
            0.0
            if boxes_value in (0, 0.0, '0', '0.0')
            else dec_to_num(parse_decimal(boxes_value))
        )
        identifiers = {
            normalize_space(row.get('sscc')),
            normalize_space(row.get('nestedInCb')),
            normalize_space(row.get('pallet')),
        }
        identifiers.discard('')
        if (
            not item_no
            or not identifiers
            or quantity is None
            or boxes is None
        ):
            continue
        for identifier in identifiers:
            group = packing_groups.setdefault(
                (item_no, identifier),
                {'quantity': 0.0, 'boxes': 0.0},
            )
            group['quantity'] += float(quantity)
            group['boxes'] += float(boxes)

    batch_groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for row in batch_rows:
        if row.get('quantityUnit') != 'boxes':
            continue
        item_no = normalize_compact_code(row.get('itemNo'))
        pallet = normalize_space(row.get('pallet'))
        if item_no and pallet:
            batch_groups.setdefault((item_no, pallet), []).append(row)

    for key, rows in batch_groups.items():
        packing = packing_groups.get(key)
        batch_quantity = sum(
            dec_to_num(parse_decimal(row.get('quantity'))) or 0.0
            for row in rows
        )
        if not packing or batch_quantity <= 0:
            warnings.append(
                f'[batch] Не удалось сопоставить batch с packing для '
                f'{key[0]}, паллета {key[1]}'
            )
            continue

        if abs(packing['quantity'] - batch_quantity) <= 0.01:
            allocated_boxes = 0.0
            for row_index, row in enumerate(rows):
                pieces = float(
                    dec_to_num(parse_decimal(row.get('quantity'))) or 0.0
                )
                if row_index == len(rows) - 1:
                    row_boxes = packing['boxes'] - allocated_boxes
                else:
                    row_boxes = round(
                        packing['boxes'] * pieces / batch_quantity,
                        6,
                    )
                    allocated_boxes += row_boxes
                row['quantity'] = pieces
                row['boxes'] = float(row_boxes)
                row['quantityUnit'] = 'pieces'
            continue

        if abs(packing['boxes'] - batch_quantity) > 0.01:
            warnings.append(
                f'[batch] Количество batch ({batch_quantity:g}) не совпадает '
                f'ни с pieces ({packing["quantity"]:g}), ни с boxes '
                f'({packing["boxes"]:g}) для {key[0]}, паллета {key[1]}'
            )
            continue

        converted_total = 0.0
        for row_index, row in enumerate(rows):
            row_boxes = (
                dec_to_num(parse_decimal(row.get('quantity'))) or 0.0
            )
            if row_index == len(rows) - 1:
                pieces = packing['quantity'] - converted_total
            else:
                pieces = round(
                    packing['quantity'] * row_boxes / batch_quantity,
                    6,
                )
                converted_total += pieces
            row['quantity'] = float(pieces)
            row['boxes'] = float(row_boxes)
            row['quantityUnit'] = 'pieces'


def discover_file_entries(input_dir: str, shipment_key: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], Optional[str]]:
    manifest_path, manifest = load_manifest(input_dir)
    manifest_entries = build_manifest_entries(input_dir, manifest)

    invoice_entries: List[Dict[str, Any]] = []
    packing_entries: List[Dict[str, Any]] = []
    batch_entries: List[Dict[str, Any]] = []

    if manifest_entries:
        for entry in manifest_entries:
            role = infer_file_role(entry['saved_name'], entry)
            if role == 'inv':
                invoice_entries.append(entry)
            elif role == 'pac':
                packing_entries.append(entry)
            elif role == 'batch':
                batch_entries.append(entry)
        return sorted(invoice_entries, key=lambda x: x['path']), sorted(packing_entries, key=lambda x: x['path']), sorted(batch_entries, key=lambda x: x['path']), manifest_path

    inv_re = re.compile(rf'^moroccanoil-inv-{re.escape(shipment_key)}-\d+\.(pdf|xlsx|xls)$', re.I)
    pac_re = re.compile(rf'^moroccanoil-pac-{re.escape(shipment_key)}-\d+\.(pdf|xlsx|xls)$', re.I)
    batch_re = re.compile(rf'^moroccanoil-batch-{re.escape(shipment_key)}-\d+\.(xlsx|xls)$', re.I)

    for root, _, files in os.walk(input_dir):
        for file_name in files:
            lower = file_name.lower()
            full_path = os.path.join(root, file_name)
            base_entry = {
                'path': full_path,
                'saved_name': file_name,
                'original_name': file_name,
                'doc_type': None,
                'meta': {},
            }

            if inv_re.match(file_name):
                invoice_entries.append(base_entry)
                continue
            if pac_re.match(file_name):
                packing_entries.append(base_entry)
                continue
            if batch_re.match(file_name):
                batch_entries.append(base_entry)
                continue

            if shipment_key in file_name:
                if lower.endswith('.pdf'):
                    if 'pack' in lower or re.search(r'\bpac\b', lower):
                        packing_entries.append(base_entry)
                        continue
                    if (
                        lower.startswith(shipment_key.lower())
                        or 'invoice' in lower
                    ):
                        invoice_entries.append(base_entry)
                        continue
                if lower.endswith(('.xlsx', '.xls')):
                    if 'batch' in lower:
                        batch_entries.append(base_entry)
                        continue
                    if 'invoice' in lower or re.search(rf'^.*{re.escape(shipment_key)}.*inv', lower):
                        invoice_entries.append(base_entry)
                        continue
                    if 'pack' in lower or re.search(rf'^.*{re.escape(shipment_key)}.*pac', lower):
                        packing_entries.append(base_entry)
                        continue

    return sorted(invoice_entries, key=lambda x: x['path']), sorted(packing_entries, key=lambda x: x['path']), sorted(batch_entries, key=lambda x: x['path']), manifest_path


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Parse MOROCCANOIL invoice + packing bundle')
    parser.add_argument('--shipment-key', required=True)
    parser.add_argument('--input-dir', required=True)
    parser.add_argument('--pretty', action='store_true')
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    shipment_key = normalize_space(args.shipment_key)
    input_dir = normalize_space(args.input_dir)

    invoice_entries, packing_entries, batch_entries, manifest_path = discover_file_entries(input_dir, shipment_key)

    if not invoice_entries:
        sys.stderr.write(f'Не найден ни один invoice файл для поставки {shipment_key}\n')
        return 2
    if not packing_entries:
        sys.stderr.write(f'Не найден ни один packing файл для поставки {shipment_key}\n')
        return 2

    warnings: List[str] = []
    invoice_rows: List[Dict[str, Any]] = []
    invoice_numbers: List[str] = []
    for entry in invoice_entries:
        if entry['path'].lower().endswith('.pdf'):
            invoice_no, rows = parse_invoice_pdf(entry, shipment_key, warnings)
        else:
            invoice_no, rows = parse_invoice_xlsx_strict(entry, shipment_key, warnings)
        if invoice_no and invoice_no not in invoice_numbers:
            invoice_numbers.append(invoice_no)
        invoice_rows.extend(rows)

    packing_rows: List[Dict[str, Any]] = []
    for entry in packing_entries:
        if entry['path'].lower().endswith('.pdf'):
            packing_rows.extend(parse_packing_pdf(entry, shipment_key, warnings))
        else:
            packing_rows.extend(parse_packing_xlsx_strict(entry, shipment_key, warnings))

    batch_rows: List[Dict[str, Any]] = []
    for entry in batch_entries:
        batch_rows.extend(parse_batch_xlsx(entry, shipment_key, warnings))
    convert_batch_box_quantities_to_pieces(
        batch_rows,
        packing_rows,
        warnings,
    )

    if not invoice_rows:
        warnings.append(f'[validate] По поставке {shipment_key} не распознано ни одной строки invoice')
    if not packing_rows:
        warnings.append(f'[validate] По поставке {shipment_key} не распознано ни одной строки packing')

    payload = {
        'shipmentKey': shipment_key,
        'invoiceNo': '; '.join(invoice_numbers),
        'invoiceDocsCount': len(invoice_entries),
        'packingDocsCount': len(packing_entries),
        'batchDocsCount': len(batch_entries),
        'invoiceFiles': [entry['path'] for entry in invoice_entries],
        'packingFiles': [entry['path'] for entry in packing_entries],
        'batchFiles': [entry['path'] for entry in batch_entries],
        'manifestPath': manifest_path,
        'manifestUsed': bool(manifest_path),
        'invoiceRows': invoice_rows,
        'packingRows': packing_rows,
        'batchRows': batch_rows,
        'warnings': warnings,
    }

    if args.pretty:
        json.dump(payload, sys.stdout, ensure_ascii=False, indent=2)
    else:
        json.dump(payload, sys.stdout, ensure_ascii=False, separators=(',', ':'))
    sys.stdout.write('\n')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())

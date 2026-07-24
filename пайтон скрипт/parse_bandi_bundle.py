#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl

try:
    import xlrd
except Exception:
    xlrd = None


INVOICE_SHEET_CANDIDATES = [
    'RUSSIA',
    'INVOICE LIST RUSSIA',
]
PACKING_SHEET_CANDIDATES = [
    'RUSSIA',
    'PACKING LIST RUSSIA',
]
MANIFEST_FILE_NAMES = ('manifest.json', 'shipment_manifest.json')

KITS_FILE_NAMES = (
    'bandi_kits.xlsx',
    'Наборы Банди для Ч З.xlsx',
    'Наборы Банди для ЧЗ.xlsx',
    'kits.xlsx',
)
KITS_DEFAULT_DIRS = (
    '/data/bandi',
    '/opt/tg_uploads/bandi',
)


def normalize_space(value: Any) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r'[^a-z0-9а-яё]+', '', str(value or '').lower())


def normalize_gtin(value: Any) -> str:
    if value is None or value == '':
        return ''
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = re.sub(r'\D', '', str(value))
    return text


def normalize_vendor(value: Any) -> str:
    return normalize_space(value).upper()


def to_number(value: Any) -> Optional[float]:
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_space(value).replace(' ', '').replace(',', '')
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_manifest(input_dir: str) -> Tuple[Optional[str], Optional[Dict[str, Any]]]:
    for file_name in MANIFEST_FILE_NAMES:
        path = os.path.join(input_dir, file_name)
        if not os.path.exists(path):
            continue
        try:
            with open(path, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
            if isinstance(data, dict):
                return path, data
        except Exception:
            pass
    return None, None


def infer_file_role(file_name: str, meta: Optional[Dict[str, Any]] = None) -> str:
    meta = meta or {}
    doc_type = normalize_space(meta.get('doc_type') or meta.get('file_type')).lower()
    if doc_type in {'inv', 'invoice'}:
        return 'inv'
    if doc_type in {'pac', 'packing'}:
        return 'pac'
    if doc_type == 'batch':
        return 'batch'

    low = file_name.lower()
    if re.search(r'(^|[-_])inv([-. _]|$)', low) or ' ci ' in f' {low} ' or 'invoice' in low:
        return 'inv'
    if re.search(r'(^|[-_])pac([-. _]|$)', low) or ' pl ' in f' {low} ' or 'packing' in low:
        return 'pac'
    if 'batch' in low:
        return 'batch'
    return 'unknown'


def infer_invoice_scope(*values: Any) -> str:
    for value in values:
        text = normalize_space(value).upper()
        if not text:
            continue
        if 'NDG' in text:
            return 'NDG'
        if 'DG' in text:
            return 'DG'
        if any(token in text for token in ('ALL', 'COMMON', 'GENERAL', 'ОБЩ')):
            return 'ALL'
    return 'UNKNOWN'


def infer_scope_from_dg_flag(value: Any) -> str:
    text = normalize_space(value).upper()
    if text == 'DG':
        return 'DG'
    if text == 'NDG':
        return 'NDG'
    return 'UNKNOWN'


def build_manifest_entries(input_dir: str, manifest: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not isinstance(manifest, dict):
        return []

    entries: List[Dict[str, Any]] = []
    for raw in manifest.get('files') or []:
        if not isinstance(raw, dict):
            continue
        saved_path = normalize_space(raw.get('saved_path'))
        saved_name = normalize_space(raw.get('saved_name'))
        path = saved_path
        if not path and saved_name:
            path = os.path.join(input_dir, saved_name)
        if not path or not os.path.exists(path):
            continue
        entries.append({
            'path': path,
            'saved_name': saved_name or os.path.basename(path),
            'original_name': normalize_space(raw.get('original_name')) or None,
            'doc_type': normalize_space(raw.get('doc_type')) or None,
            'invoice_scope': normalize_space(raw.get('invoice_scope')) or None,
            'meta': raw,
        })
    return entries


def discover_file_entries(input_dir: str, shipment_key: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Optional[str]]:
    manifest_path, manifest = load_manifest(input_dir)
    manifest_entries = build_manifest_entries(input_dir, manifest)

    invoice_entries: List[Dict[str, Any]] = []
    packing_entries: List[Dict[str, Any]] = []

    if manifest_entries:
        for entry in manifest_entries:
            role = infer_file_role(entry['saved_name'], entry)
            if role == 'inv':
                invoice_entries.append(entry)
            elif role == 'pac':
                packing_entries.append(entry)
        return sorted(invoice_entries, key=lambda x: x['path']), sorted(packing_entries, key=lambda x: x['path']), manifest_path

    inv_re = re.compile(rf'^bandi-inv-{re.escape(shipment_key)}-\d+\.(xlsx|xls)$', re.I)
    pac_re = re.compile(rf'^bandi-pac-{re.escape(shipment_key)}-\d+\.(xlsx|xls)$', re.I)

    for root, _, files in os.walk(input_dir):
        for file_name in files:
            lower = file_name.lower()
            full_path = os.path.join(root, file_name)
            original_name = file_name

            if inv_re.match(file_name):
                invoice_entries.append({
                    'path': full_path,
                    'saved_name': file_name,
                    'original_name': original_name,
                    'doc_type': 'inv',
                    'invoice_scope': infer_invoice_scope(file_name),
                    'meta': {},
                })
                continue

            if pac_re.match(file_name):
                packing_entries.append({
                    'path': full_path,
                    'saved_name': file_name,
                    'original_name': original_name,
                    'doc_type': 'pac',
                    'invoice_scope': 'PACKING',
                    'meta': {},
                })
                continue

            if shipment_key in file_name:
                if 'ci' in lower and lower.endswith(('.xlsx', '.xls')):
                    invoice_entries.append({
                        'path': full_path,
                        'saved_name': file_name,
                        'original_name': original_name,
                        'doc_type': 'inv',
                        'invoice_scope': infer_invoice_scope(file_name),
                        'meta': {},
                    })
                    continue
                if 'pl' in lower and lower.endswith(('.xlsx', '.xls')):
                    packing_entries.append({
                        'path': full_path,
                        'saved_name': file_name,
                        'original_name': original_name,
                        'doc_type': 'pac',
                        'invoice_scope': 'PACKING',
                        'meta': {},
                    })
                    continue

    return sorted(invoice_entries, key=lambda x: x['path']), sorted(packing_entries, key=lambda x: x['path']), manifest_path


def resolve_sheet_names(all_sheet_names: Sequence[str], preferred_sheets: Sequence[str]) -> List[str]:
    if not preferred_sheets:
        return list(all_sheet_names)

    exact = [name for name in preferred_sheets if name in all_sheet_names]
    if exact:
        return exact

    normalized_candidates = {normalize_header(name): name for name in preferred_sheets}
    fuzzy = []
    for name in all_sheet_names:
        if normalize_header(name) in normalized_candidates:
            fuzzy.append(name)

    if fuzzy:
        return fuzzy

    return list(all_sheet_names)


def iter_excel_rows(path: str, preferred_sheets: Sequence[str] = ()) -> Iterable[Tuple[str, List[Any]]]:
    lower = path.lower()

    if lower.endswith('.xlsx'):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ordered = resolve_sheet_names(list(wb.sheetnames), preferred_sheets)
        for sheet_name in ordered:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                yield sheet_name, list(row)

    elif lower.endswith('.xls'):
        if xlrd is None:
            raise RuntimeError('Для чтения .xls нужен пакет xlrd. Установи: /opt/moil-venv/bin/pip install xlrd')

        book = xlrd.open_workbook(path)
        ordered = resolve_sheet_names(book.sheet_names(), preferred_sheets)
        for sheet_name in ordered:
            sh = book.sheet_by_name(sheet_name)
            for rx in range(sh.nrows):
                yield sheet_name, sh.row_values(rx)

    else:
        raise RuntimeError(f'Неподдерживаемое расширение файла: {path}')


def find_header_row(rows: List[List[Any]], required_headers: Sequence[str]) -> Tuple[int, Dict[str, int]]:
    for idx, row in enumerate(rows):
        header_map: Dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            norm = normalize_header(cell)
            if norm:
                header_map[norm] = col_idx
        if all(h in header_map for h in required_headers):
            return idx, header_map
    raise RuntimeError(f'Не найдена строка заголовков. Ожидались колонки: {", ".join(required_headers)}')


def finalize_effective_scope(file_scope: str, row_dg_flag: Any) -> str:
    normalized_file_scope = infer_invoice_scope(file_scope)
    dg_scope = infer_scope_from_dg_flag(row_dg_flag)

    if normalized_file_scope in {'DG', 'NDG'}:
        return normalized_file_scope
    if dg_scope in {'DG', 'NDG'}:
        return dg_scope
    if normalized_file_scope == 'ALL':
        return 'ALL'
    return 'UNKNOWN'


def parse_invoice_file(entry: Dict[str, Any], shipment_key: str, warnings: List[str]) -> List[Dict[str, Any]]:
    path = entry['path']
    file_name = os.path.basename(path)
    original_name = entry.get('original_name') or file_name
    file_scope = infer_invoice_scope(entry.get('invoice_scope'), original_name, file_name)

    all_rows: List[List[Any]] = [row for _, row in iter_excel_rows(path, preferred_sheets=INVOICE_SHEET_CANDIDATES)]
    required = ['no', 'barcode', 'vendorcode', 'hscode', 'dgndg', 'descriptionofgoods', 'quantity']
    try:
        header_idx, header_map = find_header_row(all_rows, required)
    except Exception as e:
        warnings.append(f'[invoice] Не удалось найти заголовки в {file_name}: {e}')
        return []

    rows: List[Dict[str, Any]] = []
    qty_idx = header_map['quantity']

    for row in all_rows[header_idx + 1:]:
        line_no = to_number(row[header_map['no']] if header_map['no'] < len(row) else None)
        gtin = normalize_gtin(row[header_map['barcode']] if header_map['barcode'] < len(row) else None)
        vendor = normalize_vendor(row[header_map['vendorcode']] if header_map['vendorcode'] < len(row) else None)
        qty = to_number(row[qty_idx] if qty_idx < len(row) else None)

        if not gtin and not vendor and line_no is None:
            continue
        if not gtin and not vendor:
            continue

        after_qty = row[qty_idx + 1: qty_idx + 6]
        numeric_after_qty = [to_number(v) for v in after_qty if to_number(v) is not None]
        unit_price = numeric_after_qty[0] if len(numeric_after_qty) >= 1 else None
        amount = numeric_after_qty[1] if len(numeric_after_qty) >= 2 else None

        hs_code = row[header_map['hscode']] if header_map['hscode'] < len(row) else None
        dg_flag = normalize_space(row[header_map['dgndg']] if header_map['dgndg'] < len(row) else None).upper() or None
        description = normalize_space(row[header_map['descriptionofgoods']] if header_map['descriptionofgoods'] < len(row) else None)
        effective_scope = finalize_effective_scope(file_scope, dg_flag)

        rows.append({
            'lineNo': int(line_no) if line_no is not None else None,
            'gtin': gtin or None,
            'vendorCode': vendor or None,
            'hsCode': normalize_space(hs_code) or None,
            'dgFlag': dg_flag,
            'description': description or None,
            'quantity': qty,
            'unitPrice': unit_price,
            'amount': amount,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
            '__fileScope': file_scope,
            '__effectiveCustomsScope': effective_scope,
        })

    if not rows:
        warnings.append(f'[invoice] Не удалось распарсить строки invoice: {file_name}')

    return rows


def extract_pallet(value: Any) -> Optional[str]:
    text = normalize_space(value)
    if not text:
        return None
    m = re.search(r'(\d+)', text)
    return m.group(1) if m else None



def get_cell(row: Sequence[Any], index: Optional[int]) -> Any:
    if index is None:
        return None
    if 0 <= index < len(row):
        return row[index]
    return None


def normalize_numeric_like_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = normalize_space(value)
    return text or None

def parse_packing_file(entry: Dict[str, Any], shipment_key: str, warnings: List[str]) -> List[Dict[str, Any]]:
    path = entry['path']
    file_name = os.path.basename(path)
    original_name = entry.get('original_name') or file_name

    all_rows: List[List[Any]] = [row for _, row in iter_excel_rows(path, preferred_sheets=PACKING_SHEET_CANDIDATES)]
    required = ['no', 'barcode', 'vendercode', 'dgndg', 'descriptionofgoods', 'lot', 'quantity', 'netweight']
    try:
        header_idx, header_map = find_header_row(all_rows, required)
    except Exception as e:
        warnings.append(f'[packing] Не удалось найти заголовки в {file_name}: {e}')
        return []

    gross_idx = header_map.get('grossweight')
    carton_idx = header_map.get('cartonno')
    pallet_idx = header_map.get('palletno')

    rows: List[Dict[str, Any]] = []
    current_pallet: Optional[str] = None
    current_carton: Optional[str] = None
    current_gross_weight: Optional[float] = None

    for row in all_rows[header_idx + 1:]:
        line_no = to_number(get_cell(row, header_map['no']))
        gtin = normalize_gtin(get_cell(row, header_map['barcode']))
        vendor = normalize_vendor(get_cell(row, header_map['vendercode']))
        lot = normalize_space(get_cell(row, header_map['lot']))
        qty = to_number(get_cell(row, header_map['quantity']))
        net_weight = to_number(get_cell(row, header_map['netweight']))

        if not gtin and not vendor and line_no is None:
            continue
        if not gtin and not vendor:
            continue

        dg_flag = normalize_space(get_cell(row, header_map['dgndg'])).upper() or None
        description = normalize_space(get_cell(row, header_map['descriptionofgoods'])) or None

        raw_gross = to_number(get_cell(row, gross_idx))
        if raw_gross is not None:
            current_gross_weight = raw_gross
        gross_weight = raw_gross if raw_gross is not None else current_gross_weight

        raw_carton = normalize_numeric_like_text(get_cell(row, carton_idx))
        if raw_carton is not None:
            current_carton = raw_carton
        carton_no = raw_carton if raw_carton is not None else current_carton

        raw_pallet = extract_pallet(get_cell(row, pallet_idx))
        if raw_pallet:
            current_pallet = raw_pallet
        pallet_no = raw_pallet if raw_pallet else current_pallet

        effective_scope = finalize_effective_scope('PACKING', dg_flag)

        rows.append({
            'lineNo': int(line_no) if line_no is not None else None,
            'gtin': gtin or None,
            'vendorCode': vendor or None,
            'dgFlag': dg_flag,
            'description': description or None,
            'lot': lot or None,
            'quantity': qty,
            'netWeight': net_weight,
            'grossWeight': gross_weight,
            'cartonNo': carton_no,
            'palletNo': pallet_no,
            '__sourceFileName': file_name,
            '__sourceOriginalName': original_name,
            '__shipmentKey': shipment_key,
            '__fileScope': 'PACKING',
            '__effectiveCustomsScope': effective_scope,
        })

    if not rows:
        warnings.append(f'[packing] Не удалось распарсить строки packing: {file_name}')

    return rows


def row_key(row: Dict[str, Any]) -> str:
    gtin = normalize_space(row.get('gtin'))
    vendor = normalize_space(row.get('vendorCode'))
    line = normalize_space(row.get('lineNo'))
    description = normalize_space(row.get('description'))
    return gtin or vendor or f'{line}|{description}'


def build_scope_views(invoice_rows: List[Dict[str, Any]], packing_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    dg_invoice_dedicated_keys = {row_key(r) for r in invoice_rows if r.get('__fileScope') == 'DG'}
    ndg_invoice_dedicated_keys = {row_key(r) for r in invoice_rows if r.get('__fileScope') == 'NDG'}
    all_invoice_keys = {row_key(r) for r in invoice_rows if r.get('__fileScope') == 'ALL'}

    dg_invoice_rows: List[Dict[str, Any]] = []
    ndg_invoice_rows: List[Dict[str, Any]] = []
    cz_invoice_rows: List[Dict[str, Any]] = []

    for row in invoice_rows:
        key = row_key(row)
        eff = row.get('__effectiveCustomsScope')
        file_scope = row.get('__fileScope')

        if eff == 'DG':
            if file_scope == 'DG' or (file_scope in {'ALL', 'UNKNOWN'} and key not in dg_invoice_dedicated_keys):
                dg_invoice_rows.append(row)
        elif eff == 'NDG':
            if file_scope == 'NDG' or (file_scope in {'ALL', 'UNKNOWN'} and key not in ndg_invoice_dedicated_keys):
                ndg_invoice_rows.append(row)

        if all_invoice_keys:
            if file_scope == 'ALL':
                cz_invoice_rows.append(row)
            elif key not in all_invoice_keys:
                cz_invoice_rows.append(row)
        else:
            cz_invoice_rows.append(row)

    dg_packing_rows = [row for row in packing_rows if row.get('__effectiveCustomsScope') == 'DG']
    ndg_packing_rows = [row for row in packing_rows if row.get('__effectiveCustomsScope') == 'NDG']
    cz_packing_rows = list(packing_rows)

    return {
        'dgInvoiceRows': dg_invoice_rows,
        'ndgInvoiceRows': ndg_invoice_rows,
        'czInvoiceRows': cz_invoice_rows,
        'dgPackingRows': dg_packing_rows,
        'ndgPackingRows': ndg_packing_rows,
        'czPackingRows': cz_packing_rows,
        'invoiceScopesPresent': sorted({row.get('__fileScope') for row in invoice_rows if row.get('__fileScope')}),
        'effectiveScopesPresent': sorted({row.get('__effectiveCustomsScope') for row in invoice_rows if row.get('__effectiveCustomsScope')}),
    }


def find_kits_file(input_dir: str, explicit_path: Optional[str]) -> Optional[str]:
    candidates: List[str] = []
    if explicit_path:
        candidates.append(explicit_path)
    if input_dir:
        for nm in KITS_FILE_NAMES:
            candidates.append(os.path.join(input_dir, nm))
    for d in KITS_DEFAULT_DIRS:
        for nm in KITS_FILE_NAMES:
            candidates.append(os.path.join(d, nm))
    for path in candidates:
        if path and os.path.isfile(path):
            return path
    return None


def load_kits_mapping(path: Optional[str], warnings: List[str]) -> List[Dict[str, Any]]:
    if not path:
        return []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        warnings.append(f'[kits] Не удалось открыть справочник наборов {path}: {e}')
        return []

    sh = wb[wb.sheetnames[0]]
    rows = list(sh.iter_rows(values_only=True))

    kits: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None

    for row in rows:
        cells = list(row) + [None] * (8 - len(row)) if len(row) < 8 else list(row)
        gtin = normalize_gtin(cells[0])
        vendor = normalize_vendor(cells[1])
        dg_flag = normalize_space(cells[2]).upper() or None
        hs_code = normalize_space(cells[3]) or None
        name_eng = normalize_space(cells[4]) or None
        name_rus = normalize_space(cells[5]) or None
        volume = normalize_space(cells[6]) or None
        kind = normalize_space(cells[7]).lower()

        if not gtin and not vendor:
            continue

        if kind in ('набор', 'kit', 'set'):
            current = {
                'kitGtin': gtin or None,
                'kitVendorCode': vendor or None,
                'kitDgFlag': dg_flag,
                'kitHsCode': hs_code,
                'kitNameEng': name_eng,
                'kitNameRus': name_rus,
                'kitVolume': volume,
                'components': [],
            }
            kits.append(current)
        elif kind in ('составляющие', 'составляющая', 'component', 'components'):
            if current is None:
                warnings.append(f'[kits] Строка-составляющая без активного набора: {vendor or gtin}')
                continue
            current['components'].append({
                'gtin': gtin or None,
                'vendorCode': vendor or None,
                'dgFlag': dg_flag,
                'hsCode': hs_code,
                'nameEng': name_eng,
                'nameRus': name_rus,
                'volume': volume,
            })

    if not kits:
        warnings.append(f'[kits] Справочник наборов прочитан, но не найдено ни одного набора: {path}')

    return kits


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Parse BANDI invoice + packing bundle')
    parser.add_argument('--shipment-key', required=True)
    parser.add_argument('--input-dir', required=True)
    parser.add_argument('--kits-file', default=None,
                        help='Путь к справочнику наборов (xlsx). Если не задан — ищется в input-dir и /data/bandi/.')
    parser.add_argument('--pretty', action='store_true')
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    shipment_key = normalize_space(args.shipment_key)
    input_dir = normalize_space(args.input_dir)

    invoice_entries, packing_entries, manifest_path = discover_file_entries(input_dir, shipment_key)

    if not invoice_entries:
        sys.stderr.write(f'Не найден ни один invoice файл bandi для поставки {shipment_key}\n')
        return 2

    if not packing_entries:
        sys.stderr.write(f'Не найден ни один packing файл bandi для поставки {shipment_key}\n')
        return 2

    warnings: List[str] = []

    invoice_rows: List[Dict[str, Any]] = []
    for entry in invoice_entries:
        invoice_rows.extend(parse_invoice_file(entry, shipment_key, warnings))

    packing_rows: List[Dict[str, Any]] = []
    for entry in packing_entries:
        packing_rows.extend(parse_packing_file(entry, shipment_key, warnings))

    scope_views = build_scope_views(invoice_rows, packing_rows)

    kits_path = find_kits_file(input_dir, args.kits_file)
    kits_mapping = load_kits_mapping(kits_path, warnings)

    kit_vendor_codes = {k['kitVendorCode'] for k in kits_mapping if k.get('kitVendorCode')}
    kit_gtins = {k['kitGtin'] for k in kits_mapping if k.get('kitGtin')}

    def mark_kit_rows(rows: List[Dict[str, Any]]) -> None:
        for r in rows:
            v = r.get('vendorCode')
            g = r.get('gtin')
            r['__isKit'] = bool(
                (v and v in kit_vendor_codes) or (g and g in kit_gtins)
            )

    mark_kit_rows(invoice_rows)
    mark_kit_rows(packing_rows)

    payload = {
        'shipmentKey': shipment_key,
        'invoiceDocsCount': len(invoice_entries),
        'packingDocsCount': len(packing_entries),
        'invoiceFiles': [entry['path'] for entry in invoice_entries],
        'packingFiles': [entry['path'] for entry in packing_entries],
        'manifestPath': manifest_path,
        'manifestUsed': bool(manifest_path),
        'kitsFile': kits_path,
        'kitsCount': len(kits_mapping),
        'kitsMapping': kits_mapping,
        'invoiceRows': invoice_rows,
        'packingRows': packing_rows,
        'dgInvoiceRows': scope_views['dgInvoiceRows'],
        'ndgInvoiceRows': scope_views['ndgInvoiceRows'],
        'czInvoiceRows': scope_views['czInvoiceRows'],
        'dgPackingRows': scope_views['dgPackingRows'],
        'ndgPackingRows': scope_views['ndgPackingRows'],
        'czPackingRows': scope_views['czPackingRows'],
        'invoiceScopesPresent': scope_views['invoiceScopesPresent'],
        'effectiveScopesPresent': scope_views['effectiveScopesPresent'],
        'warnings': warnings,
    }

    if args.pretty:
        json.dump(payload, sys.stdout, ensure_ascii=False, indent=2)
    else:
        json.dump(payload, sys.stdout, ensure_ascii=False, separators=(',', ':'))
    sys.stdout.write('\n')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())

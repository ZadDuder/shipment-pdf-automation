#!/usr/bin/env python3
"""Prepare a safe MOROCCANOIL color master-catalog SKU update.

The script is deliberately offline: it reads a Google Sheets values snapshot
and the supplier XLSX, then writes a candidate values payload plus an audit
report. Applying the payload to Google Sheets is a separate, reviewed step.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
import json
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import load_workbook


SOURCE_HEADERS = {
    "Old SKU",
    "SKU",
    "SKU RU",
    "Barcode",
    "PRODUCT DESCRIPTION",
}
MASTER_REQUIRED_HEADERS = {
    "GTIN",
    "SKU Code - 1",
    "SKU Code - 2",
    "АРТИКУЛ",
    "Description",
    "BARCODE",
}
ALIAS_HEADERS = ("Old SKU FNO", "Legacy SKU Code - 2")
SKU_ALIAS_HEADERS = (
    "SKU Code - 2",
    "SKU Code - 1",
    "Old SKU FNO",
    "Legacy SKU Code - 2",
)


def clean(value: Any) -> str:
    return str(value if value is not None else "").strip()


def normalize_code(value: Any) -> str:
    # Keep this exactly aligned with normalizeCode() in the n8n build node.
    return re.sub(r"[^A-Z0-9-]", "", clean(value), flags=re.IGNORECASE).upper()


def barcode_text(value: Any) -> str:
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\D", "", clean(value))


def load_source_mappings(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Order Form"]
        raw_headers = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
        header_positions: dict[str, int] = {}
        for index, value in enumerate(raw_headers):
            header = clean(value)
            if header and header not in header_positions:
                header_positions[header] = index
        missing_headers = sorted(SOURCE_HEADERS - set(header_positions))
        if missing_headers:
            raise ValueError(
                "Supplier workbook is missing headers: "
                + ", ".join(missing_headers)
            )

        mappings: list[dict[str, Any]] = []
        for source_row, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            2,
        ):
            old_sku = clean(values[header_positions["Old SKU"]])
            new_sku = clean(values[header_positions["SKU"]])
            if not old_sku or not new_sku:
                continue
            mappings.append(
                {
                    "source_row": source_row,
                    "old_sku": old_sku,
                    "new_sku": new_sku,
                    "sku_ru": clean(values[header_positions["SKU RU"]]),
                    "barcode": barcode_text(
                        values[header_positions["Barcode"]]
                    ),
                    "description": clean(
                        values[header_positions["PRODUCT DESCRIPTION"]]
                    ),
                }
            )
    finally:
        workbook.close()

    for key in ("old_sku", "new_sku"):
        duplicates = _duplicate_mapping_values(mappings, key, normalize_code)
        if duplicates:
            raise ValueError(f"Duplicate supplier {key}: {duplicates}")
    return mappings


def _duplicate_mapping_values(
    mappings: Iterable[dict[str, Any]],
    key: str,
    normalizer,
) -> dict[str, list[int]]:
    rows_by_value: defaultdict[str, list[int]] = defaultdict(list)
    for mapping in mappings:
        value = normalizer(mapping.get(key))
        if value:
            rows_by_value[value].append(int(mapping["source_row"]))
    return {
        value: rows
        for value, rows in sorted(rows_by_value.items())
        if len(rows) > 1
    }


def _pad_rows(values: list[list[Any]]) -> list[list[str]]:
    width = max(len(row) for row in values)
    return [
        [clean(value) for value in row] + [""] * (width - len(row))
        for row in values
    ]


def build_catalog_update(
    master_values: list[list[Any]],
    mappings: list[dict[str, Any]],
) -> tuple[list[list[str]], dict[str, Any]]:
    if not master_values:
        raise ValueError("Master catalog snapshot is empty")

    candidate = _pad_rows(master_values)
    headers = candidate[0]
    missing_headers = sorted(MASTER_REQUIRED_HEADERS - set(headers))
    if missing_headers:
        raise ValueError(
            "Master catalog is missing headers: " + ", ".join(missing_headers)
        )

    for header in ALIAS_HEADERS:
        if header not in headers:
            headers.append(header)
            for row in candidate[1:]:
                row.append("")

    positions = {header: index for index, header in enumerate(headers)}
    width = len(headers)
    for row in candidate:
        row.extend([""] * (width - len(row)))

    duplicate_source_barcodes = _duplicate_mapping_values(
        mappings,
        "barcode",
        barcode_text,
    )
    source_barcode_counts = Counter(
        barcode_text(mapping.get("barcode"))
        for mapping in mappings
        if barcode_text(mapping.get("barcode"))
    )

    rows_by_old: defaultdict[str, set[int]] = defaultdict(set)
    rows_by_new: defaultdict[str, set[int]] = defaultdict(set)
    rows_by_barcode: defaultdict[str, set[int]] = defaultdict(set)
    for row_index, row in enumerate(candidate[1:], 1):
        for header in ("SKU Code - 1", "Old SKU FNO"):
            value = normalize_code(row[positions[header]])
            if value:
                rows_by_old[value].add(row_index)
        new_sku = normalize_code(row[positions["SKU Code - 2"]])
        if new_sku:
            rows_by_new[new_sku].add(row_index)
        for header in ("GTIN", "BARCODE"):
            value = barcode_text(row[positions[header]])
            if value:
                rows_by_barcode[value].add(row_index)

    report: dict[str, Any] = {
        "source_product_rows": len(mappings),
        "matched_by_old_sku": 0,
        "matched_by_new_sku": 0,
        "matched_by_barcode": 0,
        "added_rows": 0,
        "filled_blank_barcodes": 0,
        "invalid_ru_rows": [],
        "duplicate_source_barcodes": duplicate_source_barcodes,
        "updated_master_rows": [],
        "added_source_rows": [],
    }

    for mapping in mappings:
        old_sku = normalize_code(mapping.get("old_sku"))
        new_sku = normalize_code(mapping.get("new_sku"))
        barcode = barcode_text(mapping.get("barcode"))
        source_row = int(mapping["source_row"])

        matches = set(rows_by_old.get(old_sku, set()))
        method = "old_sku"
        if not matches:
            matches = set(rows_by_new.get(new_sku, set()))
            method = "new_sku"
        if (
            not matches
            and barcode
            and source_barcode_counts[barcode] == 1
        ):
            matches = set(rows_by_barcode.get(barcode, set()))
            method = "barcode"
        if len(matches) > 1:
            sheet_rows = sorted(index + 1 for index in matches)
            raise ValueError(
                f"Ambiguous mapping at supplier row {source_row}: "
                f"master rows {sheet_rows}"
            )

        sku_ru = clean(mapping.get("sku_ru"))
        valid_ru = sku_ru not in {"", "/", "-"}
        if not valid_ru:
            report["invalid_ru_rows"].append(source_row)

        if matches:
            row_index = next(iter(matches))
            row = candidate[row_index]
            legacy_sku2 = clean(row[positions["SKU Code - 2"]])
            if (
                legacy_sku2
                and normalize_code(legacy_sku2) != new_sku
                and not row[positions["Legacy SKU Code - 2"]]
            ):
                row[positions["Legacy SKU Code - 2"]] = legacy_sku2
            row[positions["SKU Code - 2"]] = clean(mapping["new_sku"])
            row[positions["Old SKU FNO"]] = clean(mapping["old_sku"])
            if valid_ru:
                row[positions["АРТИКУЛ"]] = sku_ru

            barcode_was_filled = False
            if barcode and source_barcode_counts[barcode] == 1:
                for header in ("GTIN", "BARCODE"):
                    if not barcode_text(row[positions[header]]):
                        row[positions[header]] = barcode
                        barcode_was_filled = True
            if barcode_was_filled:
                report["filled_blank_barcodes"] += 1

            report[f"matched_by_{method}"] += 1
            report["updated_master_rows"].append(row_index + 1)
            rows_by_old[old_sku].add(row_index)
            rows_by_new[new_sku].add(row_index)
            if barcode:
                rows_by_barcode[barcode].add(row_index)
            continue

        row = [""] * width
        row[positions["GTIN"]] = barcode
        row[positions["SKU Code - 1"]] = clean(mapping["old_sku"])
        row[positions["SKU Code - 2"]] = clean(mapping["new_sku"])
        row[positions["АРТИКУЛ"]] = sku_ru if valid_ru else ""
        row[positions["Description"]] = clean(mapping.get("description"))
        row[positions["BARCODE"]] = barcode
        row[positions["Old SKU FNO"]] = clean(mapping["old_sku"])
        candidate.append(row)
        row_index = len(candidate) - 1
        rows_by_old[old_sku].add(row_index)
        rows_by_new[new_sku].add(row_index)
        if barcode:
            rows_by_barcode[barcode].add(row_index)
        report["added_rows"] += 1
        report["added_source_rows"].append(source_row)

    rows_by_alias: defaultdict[str, set[int]] = defaultdict(set)
    for row_index, row in enumerate(candidate[1:], 2):
        for header in SKU_ALIAS_HEADERS:
            alias = normalize_code(row[positions[header]])
            if alias:
                rows_by_alias[alias].add(row_index)
    conflicts = {
        alias: sorted(sheet_rows)
        for alias, sheet_rows in rows_by_alias.items()
        if len(sheet_rows) > 1
    }
    if conflicts:
        alias, sheet_rows = next(iter(sorted(conflicts.items())))
        raise ValueError(
            f"Alias {alias} resolves to master rows {sheet_rows}"
        )

    report["candidate_rows_including_header"] = len(candidate)
    report["candidate_columns"] = len(headers)
    report["unique_sku_aliases"] = len(rows_by_alias)
    return candidate, report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source_xlsx", type=Path)
    parser.add_argument("master_snapshot", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    snapshot = json.loads(args.master_snapshot.read_text(encoding="utf-8"))
    values = snapshot.get("values") if isinstance(snapshot, dict) else snapshot
    if not isinstance(values, list):
        raise ValueError("Master snapshot does not contain a values array")

    mappings = load_source_mappings(args.source_xlsx)
    candidate, report = build_catalog_update(values, mappings)
    args.output.write_text(
        json.dumps({"values": candidate}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    args.report.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
